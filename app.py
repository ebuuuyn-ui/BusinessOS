import os
import sqlite3
import unicodedata
import csv
import calendar
from io import BytesIO, StringIO
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from flask import Flask, flash, make_response, redirect, render_template, request, send_file, url_for
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import event, inspect, text
from sqlalchemy.engine import Engine


db = SQLAlchemy()

ORDER_STATUSES = [
    "Bekliyor",
    "Onaylandı",
    "Üretimde",
    "Paketleniyor",
    "Sevkiyat Bekliyor",
    "Sevk Edildi",
    "Teslim Edildi",
    "İptal Edildi",
]
ORDER_TYPES = ["Satış", "Satın Alma"]
FINANCIAL_ORDER_STATUSES = ["Sevk Edildi", "Teslim Edildi"]
EXPENSE_CATEGORIES = ["Fatura Ödemeleri", "Yemek", "Ulaşım", "Kira", "Personel", "Vergi / Harç", "Bakım / Onarım", "Ofis Giderleri", "Kargo / Nakliye", "Pazarlama", "Diğer"]
PAYMENT_METHODS = ["Nakit", "Kredi Kartı"]
COLLECTION_PAYMENT_METHODS = ["Nakit", "Çek", "Banka"]
ACCOUNT_PAYMENT_METHODS = COLLECTION_PAYMENT_METHODS + ["Kredi Kartı"]
CARD_OWNER_TYPES = ["Kendi Kartımız", "Müşteri Kartı"]
CHECK_STATUSES = ["Bekliyor", "Tahsil Edildi", "Ödendi", "İade Edildi", "Karşılıksız"]


def normalize_search_text(value):
    if value is None:
        return ""
    text_value = str(value).translate(str.maketrans({"ı": "i", "İ": "I"})).casefold()
    return "".join(character for character in unicodedata.normalize("NFD", text_value) if unicodedata.category(character) != "Mn")


@event.listens_for(Engine, "connect")
def enable_sqlite_foreign_keys(dbapi_connection, _connection_record):
    if dbapi_connection.__class__.__module__.startswith("sqlite3"):
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()
        dbapi_connection.create_function("normalize_tr", 1, normalize_search_text, deterministic=True)


class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(160), nullable=False, index=True)
    code = db.Column(db.String(80), unique=True, index=True)
    contact_name = db.Column(db.String(120))
    phone = db.Column(db.String(40))
    mobile = db.Column(db.String(40))
    email = db.Column(db.String(160))
    city = db.Column(db.String(100))
    address = db.Column(db.Text)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    orders = db.relationship("Order", back_populates="customer", lazy="dynamic")
    account_transactions = db.relationship("AccountTransaction", back_populates="customer", cascade="all, delete-orphan", order_by="AccountTransaction.transaction_date")

    @property
    def balance(self):
        order_balance = sum((order.total_amount if order.order_type == "Satış" else -order.total_amount) for order in self.orders.filter(Order.status.in_(FINANCIAL_ORDER_STATUSES)).all())
        manual_balance = sum((transaction.debit or 0) - (transaction.credit or 0) for transaction in self.account_transactions)
        return order_balance + manual_balance


class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(160), nullable=False, index=True)
    code = db.Column(db.String(80), unique=True)
    description = db.Column(db.Text)
    special_code = db.Column(db.String(160))
    group_name = db.Column(db.String(160), index=True)
    default_variant = db.Column(db.String(120))
    unit = db.Column(db.String(30), default="Adet", nullable=False)
    unit_price = db.Column(db.Numeric(12, 2), default=0, nullable=False)
    purchase_price = db.Column(db.Numeric(12, 2), default=0, nullable=False)
    include_in_catalog = db.Column(db.Boolean, default=False, nullable=False, index=True)
    include_in_price_list = db.Column(db.Boolean, default=False, nullable=False, index=True)
    active = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class Order(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_no = db.Column(db.String(30), unique=True, nullable=False, index=True)
    order_type = db.Column(db.String(30), default="Satış", nullable=False, index=True)
    source_order_id = db.Column(db.Integer, nullable=True, index=True)
    customer_id = db.Column(db.Integer, db.ForeignKey("customer.id"), nullable=False)
    order_date = db.Column(db.Date, default=date.today, nullable=False)
    delivery_date = db.Column(db.Date)
    status = db.Column(db.String(40), default="Bekliyor", nullable=False, index=True)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    customer = db.relationship("Customer", back_populates="orders")
    items = db.relationship("OrderItem", back_populates="order", cascade="all, delete-orphan", order_by="OrderItem.id")
    history = db.relationship("OrderHistory", back_populates="order", cascade="all, delete-orphan", order_by="OrderHistory.created_at.desc()")

    @property
    def total_quantity(self):
        return sum(item.quantity for item in self.items)

    @property
    def net_amount(self):
        return sum((item.net_amount for item in self.items), Decimal("0"))

    @property
    def vat_amount(self):
        return sum(item.vat_amount for item in self.items)

    @property
    def total_amount(self):
        return self.net_amount + self.vat_amount


class OrderItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey("order.id", ondelete="CASCADE"), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=True)
    source_order_item_id = db.Column(db.Integer, nullable=True, index=True)
    product_name = db.Column(db.String(160), nullable=False)
    description = db.Column(db.Text)
    variant = db.Column(db.String(120))
    detail_2 = db.Column(db.String(160))
    detail_3 = db.Column(db.String(160))
    quantity = db.Column(db.Integer, nullable=False)
    unit = db.Column(db.String(30), default="Adet", nullable=False)
    unit_price = db.Column(db.Numeric(12, 2), default=0, nullable=False)
    discount_rate = db.Column(db.Numeric(5, 2), default=0, nullable=False)
    # Satış anındaki birim maliyet. Ürün kartındaki alış fiyatı sonradan
    # değişse bile geçmiş dönem kârlılığı bu değer sayesinde değişmez.
    cost_unit_price = db.Column(db.Numeric(12, 2), default=0, nullable=False)
    vat_rate = db.Column(db.Numeric(5, 2), default=10, nullable=False)
    note = db.Column(db.Text)
    order = db.relationship("Order", back_populates="items")
    product = db.relationship("Product")

    @property
    def net_amount(self):
        gross = (self.unit_price or Decimal("0")) * self.quantity
        discount = max(Decimal("0"), min(self.discount_rate or Decimal("0"), Decimal("100")))
        return gross * (Decimal("100") - discount) / Decimal("100")

    @property
    def vat_amount(self):
        return self.net_amount * (self.vat_rate or Decimal("0")) / Decimal("100")

    @property
    def total_amount(self):
        return self.net_amount + self.vat_amount

    @property
    def cost_amount(self):
        return (self.cost_unit_price or Decimal("0")) * self.quantity


class OrderHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey("order.id", ondelete="CASCADE"), nullable=False)
    status = db.Column(db.String(40), nullable=False)
    note = db.Column(db.String(240))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    order = db.relationship("Order", back_populates="history")


class AccountTransaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey("customer.id", ondelete="CASCADE"), nullable=False, index=True)
    transaction_date = db.Column(db.Date, default=date.today, nullable=False, index=True)
    transaction_type = db.Column(db.String(40), nullable=False)
    reference_no = db.Column(db.String(80))
    description = db.Column(db.String(240), nullable=False)
    debit = db.Column(db.Numeric(14, 2), default=0, nullable=False)
    credit = db.Column(db.Numeric(14, 2), default=0, nullable=False)
    payment_method = db.Column(db.String(30))
    check_no = db.Column(db.String(80))
    check_bank = db.Column(db.String(120))
    check_due_date = db.Column(db.Date, index=True)
    check_status = db.Column(db.String(40))
    card_installments = db.Column(db.Integer)
    card_owner_type = db.Column(db.String(40))
    card_customer_id = db.Column(db.Integer)
    card_customer_name = db.Column(db.String(160))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    customer = db.relationship("Customer", back_populates="account_transactions")


class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    expense_date = db.Column(db.Date, default=date.today, nullable=False, index=True)
    category = db.Column(db.String(80), nullable=False, index=True)
    document_no = db.Column(db.String(80))
    payee = db.Column(db.String(160))
    description = db.Column(db.String(240), nullable=False)
    payment_method = db.Column(db.String(40), nullable=False)
    amount = db.Column(db.Numeric(14, 2), nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class RecurringExpense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(240), nullable=False)
    category = db.Column(db.String(80), nullable=False, index=True)
    payee = db.Column(db.String(160))
    payment_method = db.Column(db.String(40), nullable=False)
    amount = db.Column(db.Numeric(14, 2), nullable=False)
    payment_day = db.Column(db.Integer, nullable=False)
    last_recorded_month = db.Column(db.String(7))
    active = db.Column(db.Boolean, default=True, nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    def due_date_for(self, target_date):
        last_day = calendar.monthrange(target_date.year, target_date.month)[1]
        return date(target_date.year, target_date.month, min(self.payment_day, last_day))

    def is_recorded_for(self, target_date):
        return self.last_recorded_month == target_date.strftime("%Y-%m")


class CashMovement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    movement_date = db.Column(db.Date, default=date.today, nullable=False, index=True)
    movement_type = db.Column(db.String(20), nullable=False)
    description = db.Column(db.String(240), nullable=False)
    amount = db.Column(db.Numeric(14, 2), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class PersonalMonth(db.Model):
    month = db.Column(db.String(7), primary_key=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    entries = db.relationship("PersonalPayment", back_populates="month_record", cascade="all, delete-orphan", order_by="PersonalPayment.sort_order")


class PersonalPayment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    month = db.Column(db.String(7), db.ForeignKey("personal_month.month", ondelete="CASCADE"), nullable=False, index=True)
    kind = db.Column(db.String(20), default="other", nullable=False)
    name = db.Column(db.String(160), nullable=False)
    credit_limit = db.Column(db.Numeric(14, 2))
    debt = db.Column(db.Numeric(14, 2))
    minimum_payment = db.Column(db.Numeric(14, 2))
    payment = db.Column(db.Numeric(14, 2))
    available_limit = db.Column(db.Numeric(14, 2))
    remaining_debt = db.Column(db.Numeric(14, 2))
    due_day = db.Column(db.Integer)
    sort_order = db.Column(db.Integer, default=0, nullable=False)
    month_record = db.relationship("PersonalMonth", back_populates="entries")

    @property
    def calculated_remaining(self):
        if self.kind == "card" and self.credit_limit is not None and self.available_limit is not None:
            return max(self.credit_limit - self.available_limit, Decimal("0"))
        return self.remaining_debt or Decimal("0")


class PersonalPerson(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(160), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    transactions = db.relationship("PersonalLedgerTransaction", back_populates="person", cascade="all, delete-orphan", order_by="PersonalLedgerTransaction.transaction_date")


class PersonalLedgerTransaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    person_id = db.Column(db.Integer, db.ForeignKey("personal_person.id", ondelete="CASCADE"), nullable=False, index=True)
    transaction_date = db.Column(db.Date)
    description = db.Column(db.String(240), default="", nullable=False)
    sent_amount = db.Column(db.Numeric(14, 2))
    received_amount = db.Column(db.Numeric(14, 2))
    sort_order = db.Column(db.Integer, default=0, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    person = db.relationship("PersonalPerson", back_populates="transactions")


def parse_date(value):
    return datetime.strptime(value, "%Y-%m-%d").date() if value else None


def parse_money(value):
    try:
        normalized = str(value or "0").strip().replace("₺", "").replace(" ", "")
        if "," in normalized:
            normalized = normalized.replace(".", "").replace(",", ".")
        elif "." in normalized:
            parts = normalized.split(".")
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3):
                normalized = "".join(parts)
        return Decimal(normalized)
    except InvalidOperation:
        return Decimal("0")


def card_payment_details(payment_method, transaction_type):
    """Validate and normalize the optional credit-card payment fields."""
    if payment_method != "Kredi Kartı":
        return {}, None
    if transaction_type != "Ödeme":
        return {}, "Kredi kartı yalnızca tedarikçiye yapılan ödemelerde kullanılabilir."
    owner_type = request.form.get("card_owner_type", "")
    try:
        installments = int(request.form.get("card_installments", "0"))
    except (TypeError, ValueError):
        installments = 0
    if owner_type not in CARD_OWNER_TYPES:
        return {}, "Kartın kime ait olduğunu seçin."
    if not 1 <= installments <= 36:
        return {}, "Taksit sayısı 1 ile 36 arasında olmalıdır."
    details = {
        "card_installments": installments,
        "card_owner_type": owner_type,
        "card_customer_id": None,
        "card_customer_name": None,
    }
    if owner_type == "Müşteri Kartı":
        card_customer_id = request.form.get("card_customer_id", type=int)
        card_customer = db.session.get(Customer, card_customer_id) if card_customer_id else None
        if not card_customer:
            return {}, "Kart sahibi müşteriyi seçin."
        details["card_customer_id"] = card_customer.id
        details["card_customer_name"] = card_customer.name
    return details, None


def import_personal_finance_data(app):
    """Eski bağımsız Ödeme Takibi verilerini Business OS'a bir kez aktarır."""
    if app.config.get("TESTING") or PersonalMonth.query.first() or PersonalPerson.query.first():
        return
    source_path = os.path.expanduser("~/Library/Application Support/OdemeTakibi/odemeler.db")
    if not os.path.isfile(source_path):
        return
    source = sqlite3.connect(source_path)
    source.row_factory = sqlite3.Row
    try:
        table_names = {row[0] for row in source.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        if {"months", "entries"}.issubset(table_names):
            for row in source.execute("SELECT * FROM months ORDER BY month"):
                db.session.add(PersonalMonth(month=row["month"]))
            db.session.flush()
            for row in source.execute("SELECT * FROM entries ORDER BY month,sort_order,id"):
                db.session.add(PersonalPayment(
                    month=row["month"], kind=row["kind"], name=row["name"], credit_limit=row["credit_limit"],
                    debt=row["debt"], minimum_payment=row["minimum_payment"], payment=row["payment"],
                    available_limit=row["available_limit"], remaining_debt=row["remaining_debt"],
                    due_day=row["due_day"], sort_order=row["sort_order"],
                ))
        person_map = {}
        if "people" in table_names:
            for row in source.execute("SELECT * FROM people ORDER BY id"):
                person = PersonalPerson(name=row["name"])
                db.session.add(person)
                db.session.flush()
                person_map[row["id"]] = person.id
        if "ledger_transactions" in table_names:
            for row in source.execute("SELECT * FROM ledger_transactions ORDER BY person_id,sort_order,id"):
                if row["person_id"] in person_map:
                    db.session.add(PersonalLedgerTransaction(
                        person_id=person_map[row["person_id"]], transaction_date=parse_date(row["transaction_date"]),
                        description=row["description"], sent_amount=row["sent_amount"], received_amount=row["received_amount"],
                        sort_order=row["sort_order"],
                    ))
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise
    finally:
        source.close()


def calculate_treasury(today=None):
    today = today or date.today()
    cash_collections = sum((item.credit or 0 for item in AccountTransaction.query.filter_by(transaction_type="Tahsilat", payment_method="Nakit").all()), Decimal("0"))
    cash_payments = sum((item.debit or 0 for item in AccountTransaction.query.filter_by(transaction_type="Ödeme", payment_method="Nakit").all()), Decimal("0"))
    cash_expenses = sum((item.amount or 0 for item in Expense.query.filter_by(payment_method="Nakit").all()), Decimal("0"))
    manual_in = sum((item.amount or 0 for item in CashMovement.query.filter_by(movement_type="Giriş").all()), Decimal("0"))
    manual_out = sum((item.amount or 0 for item in CashMovement.query.filter_by(movement_type="Çıkış").all()), Decimal("0"))
    open_checks = AccountTransaction.query.filter(AccountTransaction.payment_method == "Çek", AccountTransaction.check_status == "Bekliyor").all()
    incoming_checks = sum((item.credit or 0 for item in open_checks if item.transaction_type == "Tahsilat"), Decimal("0"))
    outgoing_checks = sum((item.debit or 0 for item in open_checks if item.transaction_type == "Ödeme"), Decimal("0"))
    due_soon = [item for item in open_checks if item.check_due_date and today <= item.check_due_date <= today + timedelta(days=30)]
    overdue_checks = [item for item in open_checks if item.check_due_date and item.check_due_date < today]
    return {
        "cash_balance": cash_collections + manual_in - cash_payments - cash_expenses - manual_out,
        "cash_collections": cash_collections,
        "cash_payments": cash_payments,
        "cash_expenses": cash_expenses,
        "manual_in": manual_in,
        "manual_out": manual_out,
        "incoming_checks": incoming_checks,
        "outgoing_checks": outgoing_checks,
        "due_soon": due_soon,
        "overdue_checks": overdue_checks,
    }


def build_account_statement(customer):
    entries = []
    for order in customer.orders.filter(Order.status.in_(FINANCIAL_ORDER_STATUSES)).all():
        is_sale = order.order_type == "Satış"
        financial_events = [event for event in order.history if event.status in FINANCIAL_ORDER_STATUSES]
        financial_event = min(financial_events, key=lambda event: event.created_at, default=None)
        financial_time = financial_event.created_at if financial_event else (order.updated_at or order.created_at)
        entries.append({
            "date": financial_time.date(),
            "sort_time": financial_time,
            "reference": order.order_no,
            "description": f"{order.order_type} Siparişi · {order.status}",
            "debit": order.total_amount if is_sale else Decimal("0"),
            "credit": Decimal("0") if is_sale else order.total_amount,
            "order_id": order.id,
            "order": order,
            "transaction_id": None,
        })
    for transaction in customer.account_transactions:
        entries.append({
            "date": transaction.transaction_date,
            "sort_time": transaction.created_at,
            "reference": transaction.reference_no or "—",
            "description": transaction.description,
            "debit": transaction.debit or Decimal("0"),
            "credit": transaction.credit or Decimal("0"),
            "order_id": None,
            "order": None,
            "transaction_id": transaction.id,
            "payment_method": transaction.payment_method,
            "check_no": transaction.check_no,
            "check_due_date": transaction.check_due_date,
            "check_status": transaction.check_status,
            "card_installments": transaction.card_installments,
            "card_owner_type": transaction.card_owner_type,
            "card_customer_id": transaction.card_customer_id,
            "card_customer_name": transaction.card_customer_name,
        })
    entries.sort(key=lambda entry: (entry["date"], entry["sort_time"], entry["reference"]))
    balance = Decimal("0")
    for entry in entries:
        balance += entry["debit"] - entry["credit"]
        entry["balance"] = balance
    return entries


def calculate_customer_balances(customer_ids=None):
    customer_query = Customer.query
    if customer_ids is not None:
        customer_query = customer_query.filter(Customer.id.in_(customer_ids))
    ids = [customer.id for customer in customer_query.all()]
    balances = {customer_id: Decimal("0") for customer_id in ids}
    if not ids:
        return balances
    for order in Order.query.filter(Order.customer_id.in_(ids), Order.status.in_(FINANCIAL_ORDER_STATUSES)).all():
        balances[order.customer_id] += order.total_amount if order.order_type == "Satış" else -order.total_amount
    for transaction in AccountTransaction.query.filter(AccountTransaction.customer_id.in_(ids)).all():
        balances[transaction.customer_id] += (transaction.debit or 0) - (transaction.credit or 0)
    return balances


def procurement_summary(sales_order, converted_orders=None):
    """Return quantity-based purchasing coverage for a sales order."""
    if sales_order.order_type != "Satış":
        return None
    linked_orders = converted_orders
    if linked_orders is None:
        linked_orders = Order.query.filter_by(source_order_id=sales_order.id).order_by(Order.id).all()
    active_orders = [order for order in linked_orders if order.status != "İptal Edildi"]
    required = {item.id: item.quantity for item in sales_order.items}
    covered = {item.id: 0 for item in sales_order.items}
    legacy_items = []
    for purchase in active_orders:
        for purchase_item in purchase.items:
            if purchase_item.source_order_item_id in covered:
                covered[purchase_item.source_order_item_id] += purchase_item.quantity
            else:
                legacy_items.append(purchase_item)

    # Eski dönüştürülmüş satın almalarda satır bağlantısı bulunmadığından ürün
    # kartı/adı ve ayrıntıları üzerinden geriye dönük eşleştirme yapılır.
    def item_key(item):
        return (
            item.product_id or 0,
            normalize_search_text(item.product_name),
            normalize_search_text(item.variant),
            normalize_search_text(item.detail_2),
            normalize_search_text(item.detail_3),
            normalize_search_text(item.unit),
        )

    sale_groups = {}
    for sale_item in sales_order.items:
        sale_groups.setdefault(item_key(sale_item), []).append(sale_item)
    for purchase_item in legacy_items:
        remaining = purchase_item.quantity
        for sale_item in sale_groups.get(item_key(purchase_item), []):
            available = max(required[sale_item.id] - covered[sale_item.id], 0)
            allocated = min(remaining, available)
            covered[sale_item.id] += allocated
            remaining -= allocated
            if remaining <= 0:
                break

    required_quantity = sum(required.values())
    covered_quantity = sum(min(covered[item_id], quantity) for item_id, quantity in required.items())
    if covered_quantity <= 0:
        code, label = "none", "Satın Alma Verilmedi"
    elif covered_quantity < required_quantity:
        code, label = "partial", "Kısmi Satın Alma"
    else:
        code, label = "complete", "Satın Alma Tamam"
    return {
        "code": code,
        "label": label,
        "required_quantity": required_quantity,
        "covered_quantity": covered_quantity,
        "linked_orders": linked_orders,
        "active_orders": active_orders,
        "items": {
            item_id: {
                "required": quantity,
                "covered": min(covered[item_id], quantity),
                "remaining": max(quantity - covered[item_id], 0),
            }
            for item_id, quantity in required.items()
        },
    }


def calculate_pending_delivery_amounts(orders=None):
    """Gelecek siparişler ve açık cari bakiyeler için sipariş bazlı nakit beklentisi."""
    if orders is None:
        orders = Order.query.filter(Order.status != "İptal Edildi").all()
    balances = calculate_customer_balances({order.customer_id for order in orders})
    grouped_orders = {}
    for order in orders:
        grouped_orders.setdefault((order.customer_id, order.order_type), []).append(order)
    expected_by_order = {}
    for (customer_id, order_type), customer_orders in grouped_orders.items():
        balance = balances.get(customer_id, Decimal("0"))
        future_total = sum((order.total_amount for order in customer_orders if order.status not in FINANCIAL_ORDER_STATUSES), Decimal("0"))
        projected_total = max(future_total + balance, Decimal("0")) if order_type == "Satış" else max(future_total - balance, Decimal("0"))
        gross_total = sum((order.total_amount for order in customer_orders), Decimal("0"))
        projected_total = min(projected_total, gross_total)
        deduction = gross_total - projected_total
        customer_orders.sort(key=lambda order: (0 if order.status in FINANCIAL_ORDER_STATUSES else 1, order.order_date, order.id))
        for order in customer_orders:
            order_deduction = min(deduction, order.total_amount)
            expected_by_order[order.id] = order.total_amount - order_deduction
            deduction -= order_deduction
    return expected_by_order


def next_order_no(order_type="Satış"):
    year = date.today().year
    prefix = f"{'SS' if order_type == 'Satış' else 'SA'}-{year}-"
    latest = Order.query.filter(Order.order_no.like(f"{prefix}%")).order_by(Order.id.desc()).first()
    sequence = 1
    if latest:
        try:
            sequence = int(latest.order_no.split("-")[-1]) + 1
        except ValueError:
            sequence = Order.query.filter(Order.order_no.like(f"{prefix}%")).count() + 1
    return f"{prefix}{sequence:05d}"


def create_database_backup(app, label="automatic", target_dir=None, keep=250):
    """Create a consistent SQLite snapshot without depending on external services."""
    if db.engine.dialect.name != "sqlite":
        return None
    database_path = db.engine.url.database
    if not database_path or database_path == ":memory:" or not os.path.exists(database_path):
        return None
    backup_dir = target_dir or os.path.join(app.instance_path, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S-%f")
    backup_path = os.path.join(backup_dir, f"business_os_{label}_{timestamp}.db")
    source = sqlite3.connect(database_path)
    destination = sqlite3.connect(backup_path)
    try:
        source.backup(destination)
    finally:
        destination.close()
        source.close()
    backups = sorted(
        (os.path.join(backup_dir, name) for name in os.listdir(backup_dir) if name.endswith(".db")),
        key=os.path.getmtime,
        reverse=True,
    )
    for old_backup in backups[keep:]:
        try:
            os.remove(old_backup)
        except OSError:
            pass
    return backup_path


def ensure_scheduled_backups(app):
    """Create one hourly local snapshot and one daily copy outside the project."""
    now = datetime.now()
    backup_dir = os.path.join(app.instance_path, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    hourly_prefix = f"business_os_hourly_{now.strftime('%Y-%m-%d_%H-')}"
    if not any(name.startswith(hourly_prefix) for name in os.listdir(backup_dir)):
        create_database_backup(app, "hourly", keep=250)
    archive_dir = os.path.expanduser("~/Documents/Business OS Yedekleri")
    daily_prefix = f"business_os_daily_{now.strftime('%Y-%m-%d_')}"
    try:
        os.makedirs(archive_dir, exist_ok=True)
        if not any(name.startswith(daily_prefix) for name in os.listdir(archive_dir)):
            create_database_backup(app, "daily", target_dir=archive_dir, keep=90)
    except OSError:
        # Harici arşiv klasörü kullanılamasa bile yerel saatlik yedek devam eder.
        pass


def personal_ledger_totals(person):
    sent = sum((item.sent_amount or Decimal("0") for item in person.transactions), Decimal("0"))
    received = sum((item.received_amount or Decimal("0") for item in person.transactions), Decimal("0"))
    return sent, received, received - sent


def safe_export_name(value):
    cleaned = "".join(character if character.isalnum() or character in "-_" else "-" for character in value.strip())
    return cleaned.strip("-") or "Kisi"


def build_personal_ledger_pdf(person):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    regular_font = "/System/Library/Fonts/Supplemental/Arial.ttf"
    bold_font = "/System/Library/Fonts/Supplemental/Arial Bold.ttf"
    font_name, bold_name = "Helvetica", "Helvetica-Bold"
    if os.path.isfile(regular_font) and os.path.isfile(bold_font):
        pdfmetrics.registerFont(TTFont("BusinessArial", regular_font))
        pdfmetrics.registerFont(TTFont("BusinessArial-Bold", bold_font))
        font_name, bold_name = "BusinessArial", "BusinessArial-Bold"

    stream = BytesIO()
    document = SimpleDocTemplate(stream, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm,
                                 topMargin=15*mm, bottomMargin=15*mm, title=f"{person.name} Borç-Alacak Ekstresi")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleTR", parent=styles["Title"], fontName=bold_name, fontSize=18, leading=22, textColor=colors.HexColor("#14213d"), alignment=TA_LEFT)
    meta_style = ParagraphStyle("MetaTR", parent=styles["BodyText"], fontName=font_name, fontSize=8, textColor=colors.HexColor("#6b7280"))
    cell_style = ParagraphStyle("CellTR", parent=styles["BodyText"], fontName=font_name, fontSize=8, leading=10)
    cell_right = ParagraphStyle("CellRightTR", parent=cell_style, alignment=TA_RIGHT)
    cell_center = ParagraphStyle("CellCenterTR", parent=cell_style, alignment=TA_CENTER)
    header_left = ParagraphStyle("HeaderLeftTR", parent=cell_style, fontName=bold_name, textColor=colors.white)
    header_right = ParagraphStyle("HeaderRightTR", parent=header_left, alignment=TA_RIGHT)
    header_center = ParagraphStyle("HeaderCenterTR", parent=header_left, alignment=TA_CENTER)
    sent, received, balance = personal_ledger_totals(person)

    def amount(value):
        return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") + " TL"

    if balance > 0:
        balance_text = f"Siz {person.name} kişisine {amount(balance)} borçlusunuz."
    elif balance < 0:
        balance_text = f"{person.name} size {amount(-balance)} borçlu."
    else:
        balance_text = f"{person.name} ile hesabınız dengede."

    story = [Paragraph(f"{person.name} - Borç / Alacak Ekstresi", title_style),
             Paragraph(f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Business OS", meta_style), Spacer(1, 5*mm)]
    summary = Table([
        ["TOPLAM GÖNDERİLEN", "TOPLAM GELEN", "NET FARK"],
        [amount(sent), amount(received), amount(abs(balance))],
        ["", "", balance_text],
    ], colWidths=[85*mm, 85*mm, 85*mm])
    summary.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#e8eef9")), ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#526074")),
        ("FONTNAME", (0,0), (-1,0), bold_name), ("FONTSIZE", (0,0), (-1,0), 8),
        ("FONTNAME", (0,1), (-1,1), bold_name), ("FONTSIZE", (0,1), (-1,1), 14),
        ("FONTNAME", (0,2), (-1,2), font_name), ("FONTSIZE", (0,2), (-1,2), 8),
        ("ALIGN", (0,0), (-1,-1), "CENTER"), ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("BOX", (0,0), (-1,-1), .5, colors.HexColor("#c9d3e3")), ("INNERGRID", (0,0), (-1,-1), .25, colors.HexColor("#dce3ec")),
        ("TOPPADDING", (0,0), (-1,-1), 6), ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    story.extend([KeepTogether(summary), Spacer(1, 6*mm)])
    rows = [[Paragraph("TARİH", header_center), Paragraph("AÇIKLAMA", header_left), Paragraph("GÖNDERİLEN PARA", header_right), Paragraph("GELEN PARA", header_right)]]
    for item in person.transactions:
        rows.append([Paragraph(item.transaction_date.strftime("%d.%m.%Y") if item.transaction_date else "-", cell_center),
                     Paragraph(item.description or "-", cell_style),
                     Paragraph(amount(item.sent_amount or Decimal("0")), cell_right),
                     Paragraph(amount(item.received_amount or Decimal("0")), cell_right)])
    rows.append(["", Paragraph("TOPLAM", ParagraphStyle("TotalLabel", parent=cell_style, fontName=bold_name)),
                 Paragraph(amount(sent), ParagraphStyle("TotalSent", parent=cell_right, fontName=bold_name)),
                 Paragraph(amount(received), ParagraphStyle("TotalReceived", parent=cell_right, fontName=bold_name))])
    ledger_table = Table(rows, colWidths=[31*mm, 129*mm, 48*mm, 48*mm], repeatRows=1)
    ledger_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#14213d")), ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), bold_name), ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.white, colors.HexColor("#f6f8fb")]),
        ("LINEBELOW", (0,0), (-1,-2), .25, colors.HexColor("#dce3ec")),
        ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#e8eef9")), ("LINEABOVE", (0,-1), (-1,-1), 1, colors.HexColor("#14213d")),
        ("TOPPADDING", (0,0), (-1,-1), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    story.append(ledger_table)

    def footer(canvas, doc):
        canvas.saveState(); canvas.setFont(font_name, 7); canvas.setFillColor(colors.HexColor("#6b7280"))
        canvas.drawString(15*mm, 8*mm, f"Business OS | {person.name} Borç-Alacak Ekstresi")
        canvas.drawRightString(landscape(A4)[0]-15*mm, 8*mm, f"Sayfa {doc.page}"); canvas.restoreState()
    document.build(story, onFirstPage=footer, onLaterPages=footer)
    stream.seek(0)
    return stream


def build_personal_ledger_xlsx(person):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Borç-Alacak Ekstresi"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A8"
    sent, received, balance = personal_ledger_totals(person)
    navy, blue, pale, green, line = "14213D", "2563EB", "E8EEF9", "14735A", "DCE3EC"
    thin = Side(style="thin", color=line)
    sheet.merge_cells("A1:D1"); sheet["A1"] = f"{person.name} - Borç / Alacak Ekstresi"
    sheet["A1"].font = Font(name="Arial", size=18, bold=True, color=navy); sheet["A1"].alignment = Alignment(horizontal="left")
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells("A2:D2"); sheet["A2"] = f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Business OS"
    sheet["A2"].font = Font(name="Arial", size=9, color="6B7280")
    sheet["A4"], sheet["B4"], sheet["C4"], sheet["D4"] = "TOPLAM GÖNDERİLEN", "TOPLAM GELEN", "NET FARK", "GÜNCEL DURUM"
    for cell in sheet[4]:
        cell.fill = PatternFill("solid", fgColor=pale); cell.font = Font(name="Arial", size=9, bold=True, color="526074"); cell.alignment = Alignment(horizontal="center")
    start_row = 8
    end_data_row = start_row + len(person.transactions) - 1
    total_row = max(end_data_row + 1, start_row)
    sheet["A5"] = f"=C{total_row}"; sheet["B5"] = f"=D{total_row}"; sheet["C5"] = "=ABS(B5-A5)"
    if balance > 0: status = f"Siz {person.name} kişisine borçlusunuz"
    elif balance < 0: status = f"{person.name} size borçlu"
    else: status = "Hesap dengede"
    sheet["D5"] = status
    for cell in sheet[5]:
        cell.font = Font(name="Arial", size=12, bold=True, color=green if cell.column != 4 else navy); cell.alignment = Alignment(horizontal="center")
    sheet.row_dimensions[5].height = 25
    headers = ["Tarih", "Açıklama", "Gönderilen Para", "Gelen Para"]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(start_row-1, column, header); cell.fill = PatternFill("solid", fgColor=navy); cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF"); cell.alignment = Alignment(horizontal="center")
    sheet.row_dimensions[start_row-1].height = 24
    for row_index, item in enumerate(person.transactions, start_row):
        sheet.cell(row_index, 1, item.transaction_date)
        sheet.cell(row_index, 2, item.description)
        sheet.cell(row_index, 3, float(item.sent_amount or 0))
        sheet.cell(row_index, 4, float(item.received_amount or 0))
        for cell in sheet[row_index]: cell.border = Border(bottom=thin); cell.font = Font(name="Arial", size=10)
    sheet.cell(total_row, 2, "TOPLAM")
    sheet.cell(total_row, 3, f"=SUM(C{start_row}:C{end_data_row})" if person.transactions else "=0")
    sheet.cell(total_row, 4, f"=SUM(D{start_row}:D{end_data_row})" if person.transactions else "=0")
    for cell in sheet[total_row]: cell.fill = PatternFill("solid", fgColor=pale); cell.font = Font(name="Arial", size=10, bold=True); cell.border = Border(top=Side(style="medium", color=navy))
    currency_format = '₺#,##0.00;[Red]-₺#,##0.00;₺-'
    for row in range(5, total_row+1):
        for column in (3,4): sheet.cell(row,column).number_format = currency_format
    for cell in (sheet["A5"], sheet["B5"], sheet["C5"]): cell.number_format = currency_format
    for row in range(start_row, end_data_row+1): sheet.cell(row,1).number_format = "dd.mm.yyyy"
    sheet.column_dimensions["A"].width = 16; sheet.column_dimensions["B"].width = 55; sheet.column_dimensions["C"].width = 23; sheet.column_dimensions["D"].width = 36
    sheet.auto_filter.ref = f"A{start_row-1}:D{max(end_data_row,start_row-1)}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True; sheet.page_setup.orientation = "landscape"; sheet.page_setup.fitToWidth = 1; sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = f"1:{start_row-1}"; sheet.print_area = f"A1:D{total_row}"
    workbook.calculation.fullCalcOnLoad = True; workbook.calculation.forceFullCalc = True; workbook.calculation.calcMode = "auto"
    stream = BytesIO(); workbook.save(stream); stream.seek(0)
    return stream


def product_cost(product_id):
    product = db.session.get(Product, product_id) if product_id else None
    return product.purchase_price if product else Decimal("0")


def order_realization_date(order):
    events = [event.created_at.date() for event in order.history if event.status in FINANCIAL_ORDER_STATUSES]
    return min(events) if events else (order.updated_at or order.created_at).date()


def latest_delivered_purchase_cost(item, cutoff_date=None):
    """Return the latest realized purchase price for the same product."""
    query = OrderItem.query.join(Order).filter(
        Order.order_type == "Satın Alma",
        Order.status.in_(FINANCIAL_ORDER_STATUSES),
    )
    if item.product_id:
        query = query.filter(OrderItem.product_id == item.product_id)
    else:
        query = query.filter(db.func.normalize_tr(OrderItem.product_name) == normalize_search_text(item.product_name))
    candidates = query.all()
    if cutoff_date:
        candidates = [candidate for candidate in candidates if order_realization_date(candidate.order) <= cutoff_date]
    if not candidates:
        return Decimal("0")
    latest = max(candidates, key=lambda candidate: (order_realization_date(candidate.order), candidate.order.id, candidate.id))
    return latest.unit_price or Decimal("0")


def effective_sales_item_cost(item, sale_date=None):
    # Kârlılıkta ürün kartındaki sabit alış fiyatı kullanılmaz. Satış tarihine
    # kadar gerçekleşmiş en son satın alma siparişinin net birim fiyatı alınır.
    return latest_delivered_purchase_cost(item, sale_date)


def create_app(test_config=None):
    app = Flask(__name__, instance_relative_config=True)
    os.makedirs(app.instance_path, exist_ok=True)
    app.config.from_mapping(
        SECRET_KEY=os.getenv("SECRET_KEY", "development-key-change-in-production"),
        SQLALCHEMY_DATABASE_URI=os.getenv("DATABASE_URL", f"sqlite:///{os.path.join(app.instance_path, 'business_os.db')}"),
        SQLALCHEMY_TRACK_MODIFICATIONS=False,
    )
    if test_config:
        app.config.update(test_config)
    db.init_app(app)

    @app.before_request
    def scheduled_database_backup():
        ensure_scheduled_backups(app)

    @app.template_filter("money")
    def money(value):
        return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    @app.get("/")
    def dashboard():
        today = date.today()
        recent_orders = Order.query.order_by(Order.created_at.desc()).limit(7).all()
        active_count = Order.query.filter(~Order.status.in_(["Teslim Edildi", "İptal Edildi"])).count()
        today_orders = Order.query.filter(Order.order_date == today, Order.status != "İptal Edildi").all()
        today_sales = [order for order in today_orders if order.order_type == "Satış"]
        today_purchases = [order for order in today_orders if order.order_type == "Satın Alma"]
        status_groups = {}
        for order_type in ORDER_TYPES:
            counts = []
            for status in ORDER_STATUSES:
                count = Order.query.filter_by(order_type=order_type, status=status).count()
                if count:
                    counts.append({"name": status, "count": count})
            status_groups[order_type] = {"items": counts, "max": max((item["count"] for item in counts), default=1), "total": sum(item["count"] for item in counts)}
        exit_orders_by_date = {}
        completed_orders = Order.query.filter(Order.status.in_(["Sevk Edildi", "Teslim Edildi"])).all()
        for order in completed_orders:
            exit_events = [event for event in order.history if event.status in ["Sevk Edildi", "Teslim Edildi"]]
            exit_date = min((event.created_at.date() for event in exit_events), default=(order.updated_at or order.created_at).date())
            exit_orders_by_date.setdefault(exit_date, []).append(order)
        week_activity = []
        for days_ago in range(6, -1, -1):
            activity_date = today - timedelta(days=days_ago)
            day_orders = exit_orders_by_date.get(activity_date, [])
            sales_count = sum(1 for order in day_orders if order.order_type == "Satış")
            purchase_count = sum(1 for order in day_orders if order.order_type == "Satın Alma")
            sales_amount = sum((order.total_amount for order in day_orders if order.order_type == "Satış"), Decimal("0"))
            purchase_amount = sum((order.total_amount for order in day_orders if order.order_type == "Satın Alma"), Decimal("0"))
            week_activity.append({"date": activity_date, "label": activity_date.strftime("%d.%m"), "sales": sales_count, "purchases": purchase_count, "total": len(day_orders), "sales_amount": sales_amount, "purchase_amount": purchase_amount})
        week_max = max((max(item["sales_amount"], item["purchase_amount"]) for item in week_activity), default=Decimal("1")) or Decimal("1")
        delivered_today = db.session.query(OrderHistory.order_id).filter(OrderHistory.status == "Teslim Edildi", db.func.date(OrderHistory.created_at) == today.isoformat()).distinct().count()
        overdue_count = Order.query.filter(Order.delivery_date < today, ~Order.status.in_(["Teslim Edildi", "İptal Edildi"])).count()
        due_today_count = Order.query.filter(Order.delivery_date == today, ~Order.status.in_(["Teslim Edildi", "İptal Edildi"])).count()
        customer_balances = calculate_customer_balances()
        total_debit_balance = sum((balance for balance in customer_balances.values() if balance > 0), Decimal("0"))
        total_credit_balance = sum((-balance for balance in customer_balances.values() if balance < 0), Decimal("0"))
        month_start = today.replace(day=1)
        today_expense = sum((expense.amount for expense in Expense.query.filter(Expense.expense_date == today).all()), Decimal("0"))
        month_expense = sum((expense.amount for expense in Expense.query.filter(Expense.expense_date >= month_start, Expense.expense_date <= today).all()), Decimal("0"))
        recurring_expenses = RecurringExpense.query.filter_by(active=True).all()
        recurring_due = sorted(
            ({"expense": item, "due_date": item.due_date_for(today), "recorded": item.is_recorded_for(today)} for item in recurring_expenses),
            key=lambda item: item["due_date"],
        )
        cashflow_orders = Order.query.filter(Order.status != "İptal Edildi").all()
        pending_expected = calculate_pending_delivery_amounts(cashflow_orders)
        pending_delivery_sales = [order for order in cashflow_orders if order.order_type == "Satış" and pending_expected.get(order.id, 0) > 0]
        pending_delivery_purchases = [order for order in cashflow_orders if order.order_type == "Satın Alma" and pending_expected.get(order.id, 0) > 0]
        treasury = calculate_treasury(today)
        month_end = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
        forecast_sales_orders = Order.query.filter(
            Order.order_type == "Satış",
            ~Order.status.in_(FINANCIAL_ORDER_STATUSES + ["İptal Edildi"]),
            db.or_(Order.delivery_date.is_(None), Order.delivery_date <= month_end),
        ).all()
        forecast_sales_by_customer = {}
        for order in forecast_sales_orders:
            forecast_sales_by_customer.setdefault(order.customer_id, Decimal("0"))
            forecast_sales_by_customer[order.customer_id] += order.total_amount
        forecast_sales_amount = sum(
            (max(total + min(customer_balances.get(customer_id, Decimal("0")), Decimal("0")), Decimal("0")) for customer_id, total in forecast_sales_by_customer.items()),
            Decimal("0"),
        )
        month_checks = AccountTransaction.query.filter(
            AccountTransaction.payment_method == "Çek",
            AccountTransaction.check_status == "Bekliyor",
            AccountTransaction.check_due_date >= today,
            AccountTransaction.check_due_date <= month_end,
        ).all()
        forecast_incoming_checks = sum((item.credit or Decimal("0") for item in month_checks if item.transaction_type == "Tahsilat"), Decimal("0"))
        forecast_outgoing_checks = sum((item.debit or Decimal("0") for item in month_checks if item.transaction_type == "Ödeme"), Decimal("0"))
        forecast_purchase_orders = Order.query.filter(
            Order.order_type == "Satın Alma",
            ~Order.status.in_(FINANCIAL_ORDER_STATUSES + ["İptal Edildi"]),
            db.or_(Order.delivery_date.is_(None), Order.delivery_date <= month_end),
        ).all()
        forecast_purchases_by_supplier = {}
        for order in forecast_purchase_orders:
            forecast_purchases_by_supplier.setdefault(order.customer_id, Decimal("0"))
            forecast_purchases_by_supplier[order.customer_id] += order.total_amount
        supplier_ids = {customer_id for (customer_id,) in db.session.query(Order.customer_id).filter(Order.order_type == "Satın Alma").distinct().all()}
        forecast_supplier_debt = sum(
            (max(forecast_purchases_by_supplier.get(customer_id, Decimal("0")) - customer_balances.get(customer_id, Decimal("0")), Decimal("0")) for customer_id in supplier_ids),
            Decimal("0"),
        )
        forecast_recurring_expenses = sum((item["expense"].amount for item in recurring_due if not item["recorded"]), Decimal("0"))
        forecast_net = treasury["cash_balance"] + forecast_sales_amount + forecast_incoming_checks - forecast_supplier_debt - forecast_recurring_expenses - forecast_outgoing_checks
        return render_template(
            "dashboard.html",
            today=today,
            recent_orders=recent_orders,
            customer_count=Customer.query.count(),
            product_count=Product.query.filter_by(active=True).count(),
            active_count=active_count,
            completed_count=Order.query.filter_by(status="Teslim Edildi").count(),
            today_sales_count=len(today_sales),
            today_sales_amount=sum((order.total_amount for order in today_sales), Decimal("0")),
            today_purchase_count=len(today_purchases),
            today_purchase_amount=sum((order.total_amount for order in today_purchases), Decimal("0")),
            today_quantity=sum(order.total_quantity for order in today_orders),
            delivered_today=delivered_today,
            due_today_count=due_today_count,
            overdue_count=overdue_count,
            total_debit_balance=total_debit_balance,
            total_credit_balance=total_credit_balance,
            net_account_balance=total_debit_balance - total_credit_balance,
            debit_customer_count=sum(1 for balance in customer_balances.values() if balance > 0),
            credit_customer_count=sum(1 for balance in customer_balances.values() if balance < 0),
            today_expense=today_expense,
            month_expense=month_expense,
            recurring_due=recurring_due,
            recurring_pending_count=sum(1 for item in recurring_due if not item["recorded"]),
            recurring_pending_amount=sum((item["expense"].amount for item in recurring_due if not item["recorded"]), Decimal("0")),
            cash_forecast={
                "month_end": month_end,
                "cash": treasury["cash_balance"],
                "sales": forecast_sales_amount,
                "incoming_checks": forecast_incoming_checks,
                "supplier_debt": forecast_supplier_debt,
                "recurring_expenses": forecast_recurring_expenses,
                "outgoing_checks": forecast_outgoing_checks,
                "net": forecast_net,
            },
            pending_delivery_sales_count=len(pending_delivery_sales),
            pending_delivery_sales_amount=sum((pending_expected[order.id] for order in pending_delivery_sales), Decimal("0")),
            pending_delivery_purchases_count=len(pending_delivery_purchases),
            pending_delivery_purchases_amount=sum((pending_expected[order.id] for order in pending_delivery_purchases), Decimal("0")),
            treasury=treasury,
            status_groups=status_groups,
            week_activity=week_activity,
            week_max=week_max,
        )

    @app.get("/yedekler")
    def backups():
        backup_dir = os.path.join(app.instance_path, "backups")
        os.makedirs(backup_dir, exist_ok=True)
        files = []
        backup_names = [item for item in os.listdir(backup_dir) if item.endswith(".db")]
        backup_names.sort(key=lambda item: os.path.getmtime(os.path.join(backup_dir, item)), reverse=True)
        for name in backup_names[:80]:
            path = os.path.join(backup_dir, name)
            files.append({"name": name, "size": os.path.getsize(path), "modified": datetime.fromtimestamp(os.path.getmtime(path))})
        archive_dir = os.path.expanduser("~/Documents/Business OS Yedekleri")
        return render_template("backups.html", files=files, archive_dir=archive_dir)

    @app.post("/yedekler/olustur")
    def create_manual_backup():
        backup_path = create_database_backup(app, "manual")
        archive_dir = os.path.expanduser("~/Documents/Business OS Yedekleri")
        create_database_backup(app, "manual", target_dir=archive_dir, keep=90)
        flash("Yeni yedek oluşturuldu ve Belgeler klasörüne de kopyalandı.", "success")
        return redirect(url_for("backups"))

    @app.get("/yedekler/guncel-indir")
    def download_current_backup():
        backup_path = create_database_backup(app, "download")
        return send_file(backup_path, as_attachment=True, download_name=f"Business-OS-Yedek-{date.today().isoformat()}.db")

    @app.get("/yedekler/<path:filename>/indir")
    def download_backup(filename):
        safe_name = os.path.basename(filename)
        backup_dir = os.path.realpath(os.path.join(app.instance_path, "backups"))
        backup_path = os.path.realpath(os.path.join(backup_dir, safe_name))
        if not backup_path.startswith(backup_dir + os.sep) or not os.path.isfile(backup_path):
            return "Yedek bulunamadı", 404
        return send_file(backup_path, as_attachment=True, download_name=safe_name)

    @app.get("/kar-zarar")
    def profitability():
        today = date.today()
        period = request.args.get("period", "month")
        if period not in {"day", "month", "year"}:
            period = "month"
        selected = parse_date(request.args.get("date")) or today
        if period == "day":
            start = end = selected
            title = selected.strftime("%d.%m.%Y")
        elif period == "year":
            start, end = date(selected.year, 1, 1), date(selected.year, 12, 31)
            title = str(selected.year)
        else:
            start = date(selected.year, selected.month, 1)
            end = date(selected.year, selected.month, calendar.monthrange(selected.year, selected.month)[1])
            title = start.strftime("%m.%Y")

        realized_sales = Order.query.filter(Order.order_type == "Satış", Order.status.in_(FINANCIAL_ORDER_STATUSES)).all()
        sales = [order for order in realized_sales if start <= order_realization_date(order) <= end]
        revenue = sum((order.net_amount for order in sales), Decimal("0"))
        cost = sum((sum((effective_sales_item_cost(item, order_realization_date(order)) * item.quantity for item in order.items), Decimal("0")) for order in sales), Decimal("0"))
        gross_profit = revenue - cost
        expenses = Expense.query.filter(Expense.expense_date >= start, Expense.expense_date <= end).all()
        expenses_total = sum((item.amount for item in expenses), Decimal("0"))
        net_profit = gross_profit - expenses_total
        markup = (gross_profit / cost * 100) if cost else None
        missing_cost_items = sum(1 for order in sales for item in order.items if not effective_sales_item_cost(item, order_realization_date(order)))

        buckets = {}
        if period == "year":
            for month in range(1, 13):
                buckets[(selected.year, month)] = {"label": f"{month:02d}.{selected.year}", "revenue": Decimal("0"), "cost": Decimal("0"), "expense": Decimal("0")}
        elif period == "month":
            for day_no in range(1, end.day + 1):
                day_value = date(selected.year, selected.month, day_no)
                buckets[day_value] = {"label": day_value.strftime("%d.%m"), "revenue": Decimal("0"), "cost": Decimal("0"), "expense": Decimal("0")}
        else:
            buckets[selected] = {"label": selected.strftime("%d.%m.%Y"), "revenue": Decimal("0"), "cost": Decimal("0"), "expense": Decimal("0")}
        for order in sales:
            realized = order_realization_date(order)
            key = (realized.year, realized.month) if period == "year" else realized
            buckets[key]["revenue"] += order.net_amount
            buckets[key]["cost"] += sum((effective_sales_item_cost(item, realized) * item.quantity for item in order.items), Decimal("0"))
        for expense in expenses:
            key = (expense.expense_date.year, expense.expense_date.month) if period == "year" else expense.expense_date
            buckets[key]["expense"] += expense.amount
        timeline = []
        for bucket in buckets.values():
            bucket["gross"] = bucket["revenue"] - bucket["cost"]
            bucket["net"] = bucket["gross"] - bucket["expense"]
            timeline.append(bucket)
        chart_max = max((max(abs(item["revenue"]), abs(item["cost"]), abs(item["net"])) for item in timeline), default=Decimal("1")) or Decimal("1")
        sales.sort(key=order_realization_date, reverse=True)
        return render_template("profitability.html", period=period, selected=selected, start=start, end=end, title=title,
            revenue=revenue, cost=cost, gross_profit=gross_profit, expenses_total=expenses_total, net_profit=net_profit,
            markup=markup, missing_cost_items=missing_cost_items, sales=sales, timeline=timeline, chart_max=chart_max,
            order_realization_date=order_realization_date, effective_sales_item_cost=effective_sales_item_cost)

    @app.get("/mali-tablolar")
    def financial_statements():
        today = date.today()
        selected = parse_date(request.args.get("date")) or today
        period = request.args.get("period", "month")
        if period not in {"month", "year"}:
            period = "month"
        if period == "year":
            period_start = date(selected.year, 1, 1)
            period_title = str(selected.year)
        else:
            period_start = date(selected.year, selected.month, 1)
            period_title = period_start.strftime("%m.%Y")
        period_end = selected

        realized_orders = Order.query.filter(Order.status.in_(FINANCIAL_ORDER_STATUSES)).all()
        period_sales = [order for order in realized_orders if order.order_type == "Satış" and period_start <= order_realization_date(order) <= period_end]
        sales_revenue = sum((order.net_amount for order in period_sales), Decimal("0"))
        sales_cost = sum((sum((effective_sales_item_cost(item, order_realization_date(order)) * item.quantity for item in order.items), Decimal("0")) for order in period_sales), Decimal("0"))
        gross_profit = sales_revenue - sales_cost
        period_expenses = Expense.query.filter(Expense.expense_date >= period_start, Expense.expense_date <= period_end).all()
        expense_by_category = {}
        for expense in period_expenses:
            expense_by_category.setdefault(expense.category, Decimal("0"))
            expense_by_category[expense.category] += expense.amount
        operating_expenses = sum(expense_by_category.values(), Decimal("0"))
        net_profit = gross_profit - operating_expenses

        balances = {customer.id: Decimal("0") for customer in Customer.query.all()}
        for order in realized_orders:
            if order_realization_date(order) <= selected:
                balances[order.customer_id] += order.total_amount if order.order_type == "Satış" else -order.total_amount
        transactions = AccountTransaction.query.filter(AccountTransaction.transaction_date <= selected).all()
        for transaction in transactions:
            balances[transaction.customer_id] += (transaction.debit or Decimal("0")) - (transaction.credit or Decimal("0"))
        receivables = sum((balance for balance in balances.values() if balance > 0), Decimal("0"))
        payables = sum((-balance for balance in balances.values() if balance < 0), Decimal("0"))

        cash_collections = sum((item.credit or Decimal("0") for item in transactions if item.transaction_type == "Tahsilat" and item.payment_method == "Nakit"), Decimal("0"))
        cash_payments = sum((item.debit or Decimal("0") for item in transactions if item.transaction_type == "Ödeme" and item.payment_method == "Nakit"), Decimal("0"))
        cash_expenses = sum((item.amount for item in Expense.query.filter(Expense.expense_date <= selected, Expense.payment_method == "Nakit").all()), Decimal("0"))
        cash_movements = CashMovement.query.filter(CashMovement.movement_date <= selected).all()
        manual_cash = sum((item.amount if item.movement_type == "Giriş" else -item.amount for item in cash_movements), Decimal("0"))
        cash_balance = cash_collections - cash_payments - cash_expenses + manual_cash
        pending_checks = [item for item in transactions if item.payment_method == "Çek" and item.check_status == "Bekliyor"]
        incoming_checks = sum((item.credit or Decimal("0") for item in pending_checks if item.transaction_type == "Tahsilat"), Decimal("0"))
        outgoing_checks = sum((item.debit or Decimal("0") for item in pending_checks if item.transaction_type == "Ödeme"), Decimal("0"))
        total_assets = cash_balance + receivables + incoming_checks
        total_liabilities = payables + outgoing_checks
        calculated_equity = total_assets - total_liabilities
        return render_template("financial_statements.html", selected=selected, period=period, period_start=period_start,
            period_end=period_end, period_title=period_title, sales_revenue=sales_revenue, sales_cost=sales_cost,
            gross_profit=gross_profit, expense_by_category=sorted(expense_by_category.items()), operating_expenses=operating_expenses,
            net_profit=net_profit, cash_balance=cash_balance, receivables=receivables, incoming_checks=incoming_checks,
            payables=payables, outgoing_checks=outgoing_checks, total_assets=total_assets,
            total_liabilities=total_liabilities, calculated_equity=calculated_equity)

    @app.route("/tahsilat-girisi", methods=["GET", "POST"])
    def quick_collection():
        customers = Customer.query.order_by(Customer.name).all()
        if request.method == "POST":
            customer_id = request.form.get("customer_id", type=int)
            customer = db.session.get(Customer, customer_id) if customer_id else None
            amount = parse_money(request.form.get("amount"))
            payment_method = request.form.get("payment_method", "")
            check_due_date = parse_date(request.form.get("check_due_date"))
            if not customer:
                flash("Lütfen listeden geçerli bir cari seçin.", "error")
            elif amount <= 0:
                flash("Tahsilat tutarı sıfırdan büyük olmalıdır.", "error")
            elif payment_method not in COLLECTION_PAYMENT_METHODS:
                flash("Tahsilat şekli olarak Nakit, Çek veya Banka seçin.", "error")
            elif payment_method == "Çek" and (not request.form.get("check_no", "").strip() or not check_due_date):
                flash("Çek numarası ve vade tarihi zorunludur.", "error")
            else:
                create_database_backup(app, "before_quick_collection")
                transaction = AccountTransaction(
                    customer=customer,
                    transaction_date=parse_date(request.form.get("transaction_date")) or date.today(),
                    transaction_type="Tahsilat",
                    reference_no=request.form.get("reference_no", "").strip(),
                    description=request.form.get("description", "").strip() or "Tahsilat",
                    debit=0,
                    credit=amount,
                    payment_method=payment_method,
                    check_no=request.form.get("check_no", "").strip() if payment_method == "Çek" else None,
                    check_bank=request.form.get("check_bank", "").strip() if payment_method == "Çek" else None,
                    check_due_date=check_due_date if payment_method == "Çek" else None,
                    check_status="Bekliyor" if payment_method == "Çek" else None,
                )
                db.session.add(transaction)
                db.session.commit()
                flash(f"{customer.name} için ₺{amount:,.2f} tahsilat kaydedildi.", "success")
                return redirect(url_for("dashboard"))
        return render_template("quick_collection.html", customers=customers, today=date.today().isoformat(), is_payment=False)

    @app.route("/odeme-girisi", methods=["GET", "POST"])
    def quick_payment():
        customers = Customer.query.order_by(Customer.name).all()
        if request.method == "POST":
            customer_id = request.form.get("customer_id", type=int)
            customer = db.session.get(Customer, customer_id) if customer_id else None
            amount = parse_money(request.form.get("amount"))
            payment_method = request.form.get("payment_method", "")
            check_due_date = parse_date(request.form.get("check_due_date"))
            card_details, card_error = card_payment_details(payment_method, "Ödeme")
            if not customer:
                flash("Lütfen listeden geçerli bir cari seçin.", "error")
            elif amount <= 0:
                flash("Ödeme tutarı sıfırdan büyük olmalıdır.", "error")
            elif payment_method not in ACCOUNT_PAYMENT_METHODS:
                flash("Ödeme şekli olarak Nakit, Çek, Banka veya Kredi Kartı seçin.", "error")
            elif payment_method == "Çek" and (not request.form.get("check_no", "").strip() or not check_due_date):
                flash("Çek numarası ve vade tarihi zorunludur.", "error")
            elif card_error:
                flash(card_error, "error")
            else:
                create_database_backup(app, "before_quick_payment")
                transaction = AccountTransaction(
                    customer=customer,
                    transaction_date=parse_date(request.form.get("transaction_date")) or date.today(),
                    transaction_type="Ödeme",
                    reference_no=request.form.get("reference_no", "").strip(),
                    description=request.form.get("description", "").strip() or "Ödeme",
                    debit=amount,
                    credit=0,
                    payment_method=payment_method,
                    check_no=request.form.get("check_no", "").strip() if payment_method == "Çek" else None,
                    check_bank=request.form.get("check_bank", "").strip() if payment_method == "Çek" else None,
                    check_due_date=check_due_date if payment_method == "Çek" else None,
                    check_status="Bekliyor" if payment_method == "Çek" else None,
                    **card_details,
                )
                db.session.add(transaction)
                db.session.commit()
                flash(f"{customer.name} için ₺{amount:,.2f} ödeme kaydedildi.", "success")
                return redirect(url_for("dashboard"))
        return render_template("quick_collection.html", customers=customers, today=date.today().isoformat(), is_payment=True)

    @app.route("/musteriler", methods=["GET", "POST"])
    def customers():
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            if not name:
                flash("Müşteri adı zorunludur.", "error")
            else:
                db.session.add(Customer(name=name, code=request.form.get("code", "").strip() or None, contact_name=request.form.get("contact_name"), phone=request.form.get("phone"), mobile=request.form.get("mobile"), email=request.form.get("email"), city=request.form.get("city"), address=request.form.get("address"), notes=request.form.get("notes")))
                db.session.commit()
                flash("Müşteri kartı oluşturuldu.", "success")
                return redirect(url_for("customers"))
        query = request.args.get("q", "").strip()
        balance_filter = request.args.get("balance", "").strip()
        view = "balances" if request.args.get("view") == "balances" or balance_filter in {"debit", "credit"} else "cards"
        balance_sort = request.args.get("sort", "amount_desc").strip()
        records = Customer.query
        if query:
            if db.engine.dialect.name == "sqlite":
                pattern = f"%{normalize_search_text(query)}%"
                records = records.filter(db.or_(
                    db.func.normalize_tr(Customer.name).like(pattern),
                    db.func.normalize_tr(Customer.code).like(pattern),
                    db.func.normalize_tr(Customer.contact_name).like(pattern),
                    db.func.normalize_tr(Customer.phone).like(pattern),
                    db.func.normalize_tr(Customer.mobile).like(pattern),
                    db.func.normalize_tr(Customer.email).like(pattern),
                    db.func.normalize_tr(Customer.city).like(pattern),
                ))
            else:
                pattern = f"%{query}%"
                records = records.filter(db.or_(Customer.name.ilike(pattern), Customer.code.ilike(pattern), Customer.contact_name.ilike(pattern), Customer.phone.ilike(pattern), Customer.mobile.ilike(pattern), Customer.email.ilike(pattern), Customer.city.ilike(pattern)))
        customer_records = records.order_by(Customer.name).all()
        balances = calculate_customer_balances([customer.id for customer in customer_records])
        if view == "balances":
            customer_records = [customer for customer in customer_records if balances.get(customer.id, Decimal("0")) != 0]
            if balance_filter == "debit":
                customer_records = [customer for customer in customer_records if balances.get(customer.id, Decimal("0")) > 0]
            elif balance_filter == "credit":
                customer_records = [customer for customer in customer_records if balances.get(customer.id, Decimal("0")) < 0]
            else:
                balance_filter = ""
            if balance_sort == "amount_asc":
                customer_records.sort(key=lambda customer: abs(balances.get(customer.id, Decimal("0"))))
            elif balance_sort == "name":
                customer_records.sort(key=lambda customer: normalize_search_text(customer.name))
            else:
                balance_sort = "amount_desc"
                customer_records.sort(key=lambda customer: abs(balances.get(customer.id, Decimal("0"))), reverse=True)
        else:
            balance_filter = ""
            balance_sort = "amount_desc"
        visible_debit_total = sum((balances[customer.id] for customer in customer_records if balances[customer.id] > 0), Decimal("0"))
        visible_credit_total = sum((-balances[customer.id] for customer in customer_records if balances[customer.id] < 0), Decimal("0"))
        return render_template("customers.html", customers=customer_records, customer_balances=balances, query=query, balance_filter=balance_filter, view=view, balance_sort=balance_sort, visible_debit_total=visible_debit_total, visible_credit_total=visible_credit_total)

    @app.route("/musteriler/aktar", methods=["GET", "POST"])
    def import_customers():
        if request.method == "POST":
            uploaded = request.files.get("file")
            if not uploaded or not uploaded.filename:
                flash("Lütfen bir Excel dosyası seçin.", "error")
                return redirect(url_for("import_customers"))
            if not uploaded.filename.lower().endswith(".xlsx"):
                flash("Yalnızca .xlsx uzantılı Excel dosyaları desteklenir.", "error")
                return redirect(url_for("import_customers"))
            try:
                create_database_backup(app, "before_customer_import")
                from openpyxl import load_workbook
                workbook = load_workbook(uploaded, read_only=True, data_only=True)
                sheet = workbook.active
                rows = sheet.iter_rows(values_only=True)
                raw_headers = next(rows, None)
                if not raw_headers:
                    raise ValueError("Dosya boş.")
                headers = {str(value or "").strip().casefold(): index for index, value in enumerate(raw_headers)}
                aliases = {
                    "code": ["cari kodu", "cari kod", "kod"],
                    "name": ["ünvan", "unvan", "müşteri adı", "firma adı"],
                    "phone": ["telefon 1", "telefon", "tel"],
                    "mobile": ["cep tel", "cep telefonu", "mobil"],
                    "email": ["e-posta", "e-posta adresi", "email"],
                    "city": ["şehir", "sehir", "il"],
                }
                columns = {key: next((headers[a] for a in names if a in headers), None) for key, names in aliases.items()}
                if columns["name"] is None:
                    raise ValueError("'Ünvan' sütunu bulunamadı.")
                added = updated = unchanged = invalid = 0
                existing_by_code = {customer.code: customer for customer in Customer.query.filter(Customer.code.isnot(None)).all()}
                for row in rows:
                    def cell(field):
                        index = columns[field]
                        return str(row[index]).strip() if index is not None and index < len(row) and row[index] is not None else ""
                    name, code = cell("name"), cell("code")
                    if not name:
                        invalid += 1
                        continue
                    values = {"name": name, "phone": cell("phone"), "mobile": cell("mobile"), "email": cell("email"), "city": cell("city")}
                    if code and code in existing_by_code:
                        customer = existing_by_code[code]
                        changed = any((getattr(customer, field) or "") != value for field, value in values.items())
                        if changed:
                            for field, value in values.items():
                                setattr(customer, field, value)
                            updated += 1
                        else:
                            unchanged += 1
                        continue
                    customer = Customer(code=code or None, notes="DİA Excel aktarımı", **values)
                    db.session.add(customer)
                    if code:
                        existing_by_code[code] = customer
                    added += 1
                db.session.commit()
                workbook.close()
                flash(f"Aktarım tamamlandı: {added} yeni müşteri eklendi, {updated} müşteri güncellendi, {unchanged} kayıt zaten günceldi, {invalid} geçersiz satır atlandı.", "success")
                return redirect(url_for("customers"))
            except Exception as exc:
                db.session.rollback()
                flash(f"Excel dosyası aktarılamadı: {exc}", "error")
        return render_template("customer_import.html")

    @app.get("/musteriler/<int:customer_id>")
    def customer_detail(customer_id):
        customer = db.get_or_404(Customer, customer_id)
        orders = customer.orders.order_by(Order.order_date.desc()).all()
        return render_template("customer_detail.html", customer=customer, orders=orders)

    @app.get("/musteriler/<int:customer_id>/cari-hesap")
    def customer_account(customer_id):
        customer = db.get_or_404(Customer, customer_id)
        all_entries = build_account_statement(customer)
        start_date = parse_date(request.args.get("start_date"))
        end_date = parse_date(request.args.get("end_date"))
        opening_balance = sum((entry["debit"] - entry["credit"] for entry in all_entries if start_date and entry["date"] < start_date), Decimal("0"))
        entries = [entry for entry in all_entries if (not start_date or entry["date"] >= start_date) and (not end_date or entry["date"] <= end_date)]
        period_debit = sum((entry["debit"] for entry in entries), Decimal("0"))
        period_credit = sum((entry["credit"] for entry in entries), Decimal("0"))
        closing_balance = opening_balance + period_debit - period_credit
        return render_template("customer_account.html", customer=customer, all_customers=Customer.query.order_by(Customer.name).all(), entries=entries, start_date=request.args.get("start_date", ""), end_date=request.args.get("end_date", ""), opening_balance=opening_balance, period_debit=period_debit, period_credit=period_credit, closing_balance=closing_balance, today=date.today().isoformat())

    @app.post("/musteriler/<int:customer_id>/cari-hesap/hareket")
    def add_account_transaction(customer_id):
        customer = db.get_or_404(Customer, customer_id)
        transaction_type = request.form.get("transaction_type", "")
        amount = parse_money(request.form.get("amount"))
        debit_types = {"Ödeme", "Borç Dekontu", "Borç Devir"}
        credit_types = {"Tahsilat", "Alacak Dekontu", "Alacak Devir"}
        payment_method = request.form.get("payment_method", "") if transaction_type in {"Tahsilat", "Ödeme"} else None
        check_due_date = parse_date(request.form.get("check_due_date"))
        card_details, card_error = card_payment_details(payment_method, transaction_type)
        if transaction_type not in debit_types | credit_types:
            flash("Lütfen geçerli bir hareket türü seçin.", "error")
        elif amount <= 0:
            flash("Tutar sıfırdan büyük olmalıdır.", "error")
        elif transaction_type == "Tahsilat" and payment_method not in COLLECTION_PAYMENT_METHODS:
            flash("Tahsilat için Nakit, Çek veya Banka seçin.", "error")
        elif transaction_type == "Ödeme" and payment_method not in ACCOUNT_PAYMENT_METHODS:
            flash("Ödeme için Nakit, Çek, Banka veya Kredi Kartı seçin.", "error")
        elif payment_method == "Çek" and (not request.form.get("check_no", "").strip() or not check_due_date):
            flash("Çek numarası ve vade tarihi zorunludur.", "error")
        elif card_error:
            flash(card_error, "error")
        else:
            create_database_backup(app, "before_account_transaction")
            transaction = AccountTransaction(customer=customer, transaction_date=parse_date(request.form.get("transaction_date")) or date.today(), transaction_type=transaction_type, reference_no=request.form.get("reference_no", "").strip(), description=request.form.get("description", "").strip() or transaction_type, debit=amount if transaction_type in debit_types else 0, credit=amount if transaction_type in credit_types else 0, payment_method=payment_method, check_no=request.form.get("check_no", "").strip() if payment_method == "Çek" else None, check_bank=request.form.get("check_bank", "").strip() if payment_method == "Çek" else None, check_due_date=check_due_date if payment_method == "Çek" else None, check_status="Bekliyor" if payment_method == "Çek" else None, **card_details)
            db.session.add(transaction)
            db.session.commit()
            flash(f"{transaction_type} hareketi cari hesaba kaydedildi.", "success")
        return redirect(url_for("customer_account", customer_id=customer.id))

    @app.route("/kasa-cek", methods=["GET"])
    def treasury():
        today = date.today()
        summary = calculate_treasury(today)
        checks = AccountTransaction.query.filter_by(payment_method="Çek").order_by(AccountTransaction.check_due_date, AccountTransaction.id).all()
        cash_movements = CashMovement.query.order_by(CashMovement.movement_date.desc(), CashMovement.id.desc()).limit(100).all()
        movements = []
        for transaction in AccountTransaction.query.filter(AccountTransaction.transaction_type.in_(["Tahsilat", "Ödeme"]), AccountTransaction.payment_method.in_(["Nakit", "Çek"])).all():
            is_incoming = transaction.transaction_type == "Tahsilat"
            is_check = transaction.payment_method == "Çek"
            movements.append({
                "date": transaction.transaction_date,
                "kind": "Çek" if is_check else "Kasa",
                "direction": ("Alınan" if is_incoming else "Verilen") if is_check else ("Giriş" if is_incoming else "Çıkış"),
                "description": transaction.description,
                "party": transaction.customer.name,
                "reference": transaction.check_no if is_check else transaction.reference_no,
                "due_date": transaction.check_due_date if is_check else None,
                "status": transaction.check_status if is_check else "Gerçekleşti",
                "amount": transaction.credit if is_incoming else transaction.debit,
                "customer_id": transaction.customer_id,
                "sort_time": transaction.created_at,
                "source": "Cari Tahsilat" if is_incoming else "Cari Ödeme",
                "manual_id": None,
            })
        for expense in Expense.query.filter_by(payment_method="Nakit").all():
            movements.append({"date": expense.expense_date, "kind": "Kasa", "direction": "Çıkış", "description": expense.description, "party": expense.payee or expense.category, "reference": expense.document_no, "due_date": None, "status": "Gerçekleşti", "amount": expense.amount, "customer_id": None, "sort_time": expense.created_at, "source": "Masraf", "manual_id": None})
        for movement in CashMovement.query.all():
            movements.append({"date": movement.movement_date, "kind": "Kasa", "direction": movement.movement_type, "description": movement.description, "party": "Kasa", "reference": None, "due_date": None, "status": "Gerçekleşti", "amount": movement.amount, "customer_id": None, "sort_time": movement.created_at, "source": "Manuel", "manual_id": movement.id})
        all_cash_movements = [movement for movement in movements if movement["kind"] == "Kasa"]
        all_cash_movements.sort(key=lambda movement: (movement["date"], movement["sort_time"]), reverse=True)
        movement_filter = request.args.get("movement", "all")
        direction_filter = request.args.get("direction", "all")
        status_filter = request.args.get("status", "all")
        query = request.args.get("q", "").strip()
        start_date = parse_date(request.args.get("start_date"))
        end_date = parse_date(request.args.get("end_date"))
        if movement_filter in {"cash", "check"}:
            expected_kind = "Kasa" if movement_filter == "cash" else "Çek"
            movements = [movement for movement in movements if movement["kind"] == expected_kind]
        else:
            movement_filter = "all"
        if direction_filter == "in":
            movements = [movement for movement in movements if movement["direction"] in {"Giriş", "Alınan"}]
        elif direction_filter == "out":
            movements = [movement for movement in movements if movement["direction"] in {"Çıkış", "Verilen"}]
        else:
            direction_filter = "all"
        if status_filter != "all":
            movements = [movement for movement in movements if movement["status"] == status_filter]
        if start_date:
            movements = [movement for movement in movements if movement["date"] >= start_date]
        if end_date:
            movements = [movement for movement in movements if movement["date"] <= end_date]
        if query:
            normalized_query = normalize_search_text(query)
            movements = [movement for movement in movements if normalized_query in normalize_search_text(" ".join(str(movement.get(field) or "") for field in ["description", "party", "reference"]))]
        movements.sort(key=lambda movement: (movement["date"], movement["sort_time"]), reverse=True)
        filtered_in = sum((movement["amount"] or 0 for movement in movements if movement["direction"] in {"Giriş", "Alınan"}), Decimal("0"))
        filtered_out = sum((movement["amount"] or 0 for movement in movements if movement["direction"] in {"Çıkış", "Verilen"}), Decimal("0"))
        return render_template("treasury.html", summary=summary, checks=checks, cash_movements=cash_movements, all_cash_movements=all_cash_movements, movements=movements, filtered_in=filtered_in, filtered_out=filtered_out, today=today.isoformat(), check_statuses=CHECK_STATUSES, movement_filter=movement_filter, direction_filter=direction_filter, status_filter=status_filter, query=query, start_date=request.args.get("start_date", ""), end_date=request.args.get("end_date", ""))

    @app.post("/kasa-cek/kasa-hareketi")
    def add_cash_movement():
        movement_type = request.form.get("movement_type", "")
        amount = parse_money(request.form.get("amount"))
        description = request.form.get("description", "").strip()
        if movement_type not in {"Giriş", "Çıkış"} or amount <= 0 or not description:
            flash("Kasa hareketinin türünü, açıklamasını ve tutarını kontrol edin.", "error")
        else:
            create_database_backup(app, "before_cash_movement")
            db.session.add(CashMovement(movement_date=parse_date(request.form.get("movement_date")) or date.today(), movement_type=movement_type, description=description, amount=amount))
            db.session.commit()
            flash("Kasa hareketi kaydedildi.", "success")
        return redirect(url_for("treasury"))

    @app.post("/kasa-cek/kasa-hareketi/<int:movement_id>/sil")
    def delete_cash_movement(movement_id):
        movement = db.get_or_404(CashMovement, movement_id)
        create_database_backup(app, "before_cash_movement_delete")
        db.session.delete(movement)
        db.session.commit()
        flash("Kasa hareketi silindi.", "success")
        return redirect(url_for("treasury"))

    @app.post("/kasa-cek/cek/<int:transaction_id>/durum")
    def update_check_status(transaction_id):
        transaction = db.get_or_404(AccountTransaction, transaction_id)
        status = request.form.get("check_status", "")
        if transaction.payment_method != "Çek" or status not in CHECK_STATUSES:
            flash("Geçersiz çek durumu.", "error")
        else:
            create_database_backup(app, "before_check_status")
            transaction.check_status = status
            db.session.commit()
            flash("Çek durumu güncellendi.", "success")
        return redirect(url_for("treasury"))

    @app.post("/musteriler/<int:customer_id>/cari-hesap/hareket/<int:transaction_id>/sil")
    def delete_account_transaction(customer_id, transaction_id):
        transaction = db.get_or_404(AccountTransaction, transaction_id)
        if transaction.customer_id != customer_id:
            return ("Geçersiz cari hareketi", 400)
        create_database_backup(app, "before_account_transaction_delete")
        db.session.delete(transaction)
        db.session.commit()
        flash("Cari hesap hareketi silindi.", "success")
        return redirect(url_for("customer_account", customer_id=customer_id))

    @app.get("/musteriler/<int:customer_id>/cari-hesap/ekstre.csv")
    def customer_account_csv(customer_id):
        customer = db.get_or_404(Customer, customer_id)
        entries = build_account_statement(customer)
        start_date = parse_date(request.args.get("start_date"))
        end_date = parse_date(request.args.get("end_date"))
        entries = [entry for entry in entries if (not start_date or entry["date"] >= start_date) and (not end_date or entry["date"] <= end_date)]
        output = StringIO()
        writer = csv.writer(output, delimiter=";")
        writer.writerow(["Cari Kodu", customer.code or "", "Cari Ünvanı", customer.name])
        writer.writerow(["Tarih", "Sipariş / Referans No", "Açıklama", "Borç", "Alacak", "Bakiye"])
        for entry in entries:
            writer.writerow([entry["date"].strftime("%d.%m.%Y"), entry["reference"], entry["description"], f"{entry['debit']:.2f}", f"{entry['credit']:.2f}", f"{entry['balance']:.2f}"])
        response = make_response("\ufeff" + output.getvalue())
        response.headers["Content-Type"] = "text/csv; charset=utf-8"
        response.headers["Content-Disposition"] = f'attachment; filename="cari-ekstre-{customer.code or customer.id}.csv"'
        return response

    @app.route("/musteriler/<int:customer_id>/duzenle", methods=["GET", "POST"])
    def edit_customer(customer_id):
        customer = db.get_or_404(Customer, customer_id)
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            code = request.form.get("code", "").strip() or None
            duplicate = Customer.query.filter(Customer.code == code, Customer.id != customer.id).first() if code else None
            if not name:
                flash("Müşteri adı zorunludur.", "error")
            elif duplicate:
                flash("Bu cari kod başka bir müşteride kullanılıyor.", "error")
            else:
                customer.name = name
                customer.code = code
                customer.contact_name = request.form.get("contact_name", "").strip()
                customer.phone = request.form.get("phone", "").strip()
                customer.mobile = request.form.get("mobile", "").strip()
                customer.email = request.form.get("email", "").strip()
                customer.city = request.form.get("city", "").strip()
                customer.address = request.form.get("address", "").strip()
                customer.notes = request.form.get("notes", "").strip()
                db.session.commit()
                flash("Müşteri bilgileri güncellendi.", "success")
                return redirect(url_for("customer_detail", customer_id=customer.id))
        return render_template("customer_edit.html", customer=customer)

    @app.post("/musteriler/<int:customer_id>/sil")
    def delete_customer(customer_id):
        customer = db.get_or_404(Customer, customer_id)
        if customer.orders.count() > 0:
            flash("Bu müşteriye ait siparişler bulunduğu için müşteri silinemedi. Sipariş geçmişinin korunması gerekir.", "error")
            return redirect(url_for("customer_detail", customer_id=customer.id))
        if customer.account_transactions:
            flash("Bu müşteriye ait cari hesap hareketleri bulunduğu için müşteri silinemedi.", "error")
            return redirect(url_for("customer_detail", customer_id=customer.id))
        create_database_backup(app, "before_customer_delete")
        name = customer.name
        db.session.delete(customer)
        db.session.commit()
        flash(f"{name} müşteri kaydı silindi.", "success")
        return redirect(url_for("customers"))

    @app.route("/urunler", methods=["GET", "POST"])
    def products():
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            if not name:
                flash("Ürün adı zorunludur.", "error")
            else:
                code = request.form.get("code", "").strip() or None
                if code and Product.query.filter_by(code=code).first():
                    flash("Bu ürün kodu zaten kullanılıyor.", "error")
                else:
                    db.session.add(Product(name=name, code=code, description=request.form.get("description"), special_code=request.form.get("special_code"), group_name=request.form.get("group_name"), default_variant=request.form.get("default_variant"), unit=request.form.get("unit") or "Adet", unit_price=parse_money(request.form.get("unit_price")), purchase_price=parse_money(request.form.get("purchase_price")), include_in_catalog=request.form.get("include_in_catalog") == "on", include_in_price_list=request.form.get("include_in_price_list") == "on"))
                    db.session.commit()
                    flash("Ürün kartı oluşturuldu.", "success")
                    return redirect(url_for("products"))
        query = request.args.get("q", "").strip()
        records = Product.query
        if query:
            if db.engine.dialect.name == "sqlite":
                pattern = f"%{normalize_search_text(query)}%"
                records = records.filter(db.or_(
                    db.func.normalize_tr(Product.name).like(pattern),
                    db.func.normalize_tr(Product.code).like(pattern),
                    db.func.normalize_tr(Product.description).like(pattern),
                    db.func.normalize_tr(Product.special_code).like(pattern),
                    db.func.normalize_tr(Product.group_name).like(pattern),
                    db.func.normalize_tr(Product.default_variant).like(pattern),
                ))
            else:
                pattern = f"%{query}%"
                records = records.filter(db.or_(Product.name.ilike(pattern), Product.code.ilike(pattern), Product.description.ilike(pattern), Product.special_code.ilike(pattern), Product.group_name.ilike(pattern), Product.default_variant.ilike(pattern)))
        return render_template("products.html", products=records.order_by(Product.name).all(), query=query, catalog_count=Product.query.filter_by(include_in_catalog=True, active=True).count(), price_list_count=Product.query.filter_by(include_in_price_list=True, active=True).count())

    @app.route("/urunler/<int:product_id>/duzenle", methods=["GET", "POST"])
    def edit_product(product_id):
        product = db.get_or_404(Product, product_id)
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            code = request.form.get("code", "").strip() or None
            duplicate = Product.query.filter(Product.code == code, Product.id != product.id).first() if code else None
            if not name:
                flash("Ürün adı zorunludur.", "error")
            elif duplicate:
                flash("Bu ürün kodu başka bir kartta kullanılıyor.", "error")
            else:
                product.name = name
                product.code = code
                product.description = request.form.get("description", "").strip()
                product.special_code = request.form.get("special_code", "").strip()
                product.group_name = request.form.get("group_name", "").strip()
                product.default_variant = request.form.get("default_variant", "").strip()
                product.unit = request.form.get("unit") or "Adet"
                product.unit_price = parse_money(request.form.get("unit_price"))
                product.purchase_price = parse_money(request.form.get("purchase_price"))
                product.include_in_catalog = request.form.get("include_in_catalog") == "on"
                product.include_in_price_list = request.form.get("include_in_price_list") == "on"
                db.session.commit()
                flash("Ürün kartı ve fiyatları güncellendi.", "success")
                return redirect(url_for("products", q=product.code or product.name))
        return render_template("product_edit.html", product=product)

    @app.post("/urunler/<int:product_id>/secim")
    def toggle_product_selection(product_id):
        product = db.get_or_404(Product, product_id)
        target = request.form.get("target")
        if target not in {"catalog", "price_list"}:
            flash("Geçersiz ürün seçimi.", "error")
            return redirect(url_for("products"))
        create_database_backup(app, "before_product_selection")
        if target == "catalog":
            product.include_in_catalog = not product.include_in_catalog
            label = "katalog"
            selected = product.include_in_catalog
        else:
            product.include_in_price_list = not product.include_in_price_list
            label = "fiyat listesi"
            selected = product.include_in_price_list
        db.session.commit()
        flash(f"{product.name} {label} seçimine {'eklendi' if selected else 'çıkarıldı'}.", "success")
        return redirect(url_for("products", q=request.form.get("q", "")))

    @app.get("/katalog")
    def product_catalog():
        selected_products = Product.query.filter_by(include_in_catalog=True, active=True).order_by(Product.group_name, Product.name).all()
        return render_template("product_catalog.html", products=selected_products, generated_at=datetime.now())

    @app.get("/fiyat-listesi")
    def product_price_list():
        selected_products = Product.query.filter_by(include_in_price_list=True, active=True).order_by(Product.group_name, Product.name).all()
        return render_template("product_price_list.html", products=selected_products, generated_at=datetime.now())

    @app.get("/fiyat-listesi/csv")
    def product_price_list_csv():
        selected_products = Product.query.filter_by(include_in_price_list=True, active=True).order_by(Product.group_name, Product.name).all()
        output = StringIO()
        output.write("\ufeff")
        writer = csv.writer(output, delimiter=";")
        writer.writerow(["Ürün Kodu", "Ürün Adı", "Ürün Grubu", "Ayrıntı", "Birim", "Satış Fiyatı (₺)"])
        for product in selected_products:
            writer.writerow([product.code or "", product.name, product.group_name or "", product.default_variant or "", product.unit, f"{product.unit_price:.2f}".replace(".", ",")])
        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "text/csv; charset=utf-8"
        response.headers["Content-Disposition"] = f"attachment; filename=fiyat-listesi-{date.today().isoformat()}.csv"
        return response

    @app.route("/urunler/aktar", methods=["GET", "POST"])
    def import_products():
        if request.method == "POST":
            uploaded = request.files.get("file")
            if not uploaded or not uploaded.filename:
                flash("Lütfen bir Excel dosyası seçin.", "error")
                return redirect(url_for("import_products"))
            if not uploaded.filename.lower().endswith(".xlsx"):
                flash("Yalnızca .xlsx uzantılı Excel dosyaları desteklenir.", "error")
                return redirect(url_for("import_products"))
            try:
                create_database_backup(app, "before_product_import")
                from openpyxl import load_workbook
                workbook = load_workbook(uploaded, read_only=True, data_only=True)
                sheet = workbook.active
                rows = sheet.iter_rows(values_only=True)
                raw_headers = next(rows, None)
                if not raw_headers:
                    raise ValueError("Dosya boş.")
                headers = {str(value or "").strip().casefold(): index for index, value in enumerate(raw_headers)}
                aliases = {
                    "code": ["kart kodu", "stok kodu", "ürün kodu", "urun kodu", "kod"],
                    "name": ["açıklama", "aciklama", "ürün adı", "urun adi", "stok adı"],
                    "special_code": ["özel kod 3", "ozel kod 3", "özel kod", "ozel kod"],
                    "group_name": ["grup kodu açıklama", "grup kodu aciklama", "grup açıklama", "ürün grubu"],
                }
                columns = {key: next((headers[alias] for alias in names if alias in headers), None) for key, names in aliases.items()}
                if columns["code"] is None or columns["name"] is None:
                    raise ValueError("'Kart Kodu' veya 'Açıklama' sütunu bulunamadı.")
                added = updated = unchanged = invalid = 0
                existing_by_code = {product.code: product for product in Product.query.filter(Product.code.isnot(None)).all()}
                for row in rows:
                    def cell(field):
                        index = columns[field]
                        return str(row[index]).strip() if index is not None and index < len(row) and row[index] is not None else ""
                    code, name = cell("code"), cell("name")
                    if not code or not name:
                        invalid += 1
                        continue
                    values = {"name": name, "special_code": cell("special_code"), "group_name": cell("group_name")}
                    if code in existing_by_code:
                        product = existing_by_code[code]
                        changed = any((getattr(product, field) or "") != value for field, value in values.items())
                        if changed:
                            for field, value in values.items():
                                setattr(product, field, value)
                            updated += 1
                        else:
                            unchanged += 1
                    else:
                        product = Product(code=code, unit="Adet", unit_price=0, active=True, **values)
                        db.session.add(product)
                        existing_by_code[code] = product
                        added += 1
                db.session.commit()
                workbook.close()
                flash(f"Ürün aktarımı tamamlandı: {added} yeni ürün eklendi, {updated} ürün güncellendi, {unchanged} kayıt zaten günceldi, {invalid} geçersiz satır atlandı.", "success")
                return redirect(url_for("products"))
            except Exception as exc:
                db.session.rollback()
                flash(f"Ürün Excel dosyası aktarılamadı: {exc}", "error")
        return render_template("product_import.html")

    @app.route("/masraflar", methods=["GET", "POST"])
    def expenses():
        if request.method == "POST":
            category = request.form.get("category", "")
            payment_method = request.form.get("payment_method", "")
            amount = parse_money(request.form.get("amount"))
            description = request.form.get("description", "").strip()
            if category not in EXPENSE_CATEGORIES:
                flash("Lütfen geçerli bir masraf kategorisi seçin.", "error")
            elif payment_method not in PAYMENT_METHODS:
                flash("Lütfen geçerli bir ödeme yöntemi seçin.", "error")
            elif not description:
                flash("Masraf açıklaması zorunludur.", "error")
            elif amount <= 0:
                flash("Masraf tutarı sıfırdan büyük olmalıdır.", "error")
            else:
                create_database_backup(app, "before_expense_add")
                expense = Expense(expense_date=parse_date(request.form.get("expense_date")) or date.today(), category=category, document_no=request.form.get("document_no", "").strip(), payee=request.form.get("payee", "").strip(), description=description, payment_method=payment_method, amount=amount)
                db.session.add(expense)
                db.session.commit()
                flash("Masraf kaydı oluşturuldu.", "success")
                return redirect(url_for("expenses"))
        query = request.args.get("q", "").strip()
        category = request.args.get("category", "").strip()
        start_date = parse_date(request.args.get("start_date"))
        end_date = parse_date(request.args.get("end_date"))
        records = Expense.query
        if query:
            pattern = f"%{normalize_search_text(query)}%"
            if db.engine.dialect.name == "sqlite":
                records = records.filter(db.or_(db.func.normalize_tr(Expense.description).like(pattern), db.func.normalize_tr(Expense.document_no).like(pattern), db.func.normalize_tr(Expense.payee).like(pattern)))
            else:
                records = records.filter(db.or_(Expense.description.ilike(f"%{query}%"), Expense.document_no.ilike(f"%{query}%"), Expense.payee.ilike(f"%{query}%")))
        if category in EXPENSE_CATEGORIES:
            records = records.filter(Expense.category == category)
        if start_date:
            records = records.filter(Expense.expense_date >= start_date)
        if end_date:
            records = records.filter(Expense.expense_date <= end_date)
        expense_records = records.order_by(Expense.expense_date.desc(), Expense.id.desc()).all()
        total = sum((expense.amount for expense in expense_records), Decimal("0"))
        today = date.today()
        recurring_expenses = RecurringExpense.query.order_by(RecurringExpense.active.desc(), RecurringExpense.payment_day, RecurringExpense.description).all()
        recurring_due = [{"expense": item, "due_date": item.due_date_for(today), "recorded": item.is_recorded_for(today)} for item in recurring_expenses]
        return render_template("expenses.html", expenses=expense_records, total=total, categories=EXPENSE_CATEGORIES, payment_methods=PAYMENT_METHODS, selected_category=category, query=query, start_date=request.args.get("start_date", ""), end_date=request.args.get("end_date", ""), today=today.isoformat(), current_date=today, recurring_due=recurring_due)

    @app.post("/masraflar/aylik-plan/ekle")
    def add_recurring_expense():
        category = request.form.get("category", "")
        payment_method = request.form.get("payment_method", "")
        description = request.form.get("description", "").strip()
        amount = parse_money(request.form.get("amount"))
        try:
            payment_day = int(request.form.get("payment_day", "0"))
        except ValueError:
            payment_day = 0
        if category not in EXPENSE_CATEGORIES or payment_method not in PAYMENT_METHODS or not description or amount <= 0 or not 1 <= payment_day <= 31:
            flash("Aylık ödeme planındaki zorunlu bilgileri kontrol edin.", "error")
        else:
            create_database_backup(app, "before_recurring_expense_add")
            db.session.add(RecurringExpense(description=description, category=category, payee=request.form.get("payee", "").strip(), payment_method=payment_method, amount=amount, payment_day=payment_day))
            db.session.commit()
            flash("Aylık ödeme planı eklendi.", "success")
        return redirect(url_for("expenses") + "#aylik-odemeler")

    @app.post("/masraflar/aylik-plan/<int:plan_id>/isle")
    def record_recurring_expense(plan_id):
        plan = db.get_or_404(RecurringExpense, plan_id)
        today = date.today()
        if plan.is_recorded_for(today):
            flash("Bu ödeme bu ay zaten masraf olarak işlendi.", "error")
        else:
            create_database_backup(app, "before_recurring_expense_record")
            db.session.add(Expense(expense_date=today, category=plan.category, payee=plan.payee, description=plan.description, payment_method=plan.payment_method, amount=plan.amount))
            plan.last_recorded_month = today.strftime("%Y-%m")
            db.session.commit()
            flash("Planlı ödeme bu ayın masraf kaydına aktarıldı.", "success")
        return redirect(url_for("expenses") + "#aylik-odemeler")

    @app.post("/masraflar/aylik-plan/<int:plan_id>/sil")
    def delete_recurring_expense(plan_id):
        plan = db.get_or_404(RecurringExpense, plan_id)
        create_database_backup(app, "before_recurring_expense_delete")
        db.session.delete(plan)
        db.session.commit()
        flash("Aylık ödeme planı silindi. Önceki masraf kayıtları korundu.", "success")
        return redirect(url_for("expenses") + "#aylik-odemeler")

    @app.route("/masraflar/<int:expense_id>/duzenle", methods=["GET", "POST"])
    def edit_expense(expense_id):
        expense = db.get_or_404(Expense, expense_id)
        if request.method == "POST":
            category = request.form.get("category", "")
            payment_method = request.form.get("payment_method", "")
            amount = parse_money(request.form.get("amount"))
            description = request.form.get("description", "").strip()
            if category not in EXPENSE_CATEGORIES or payment_method not in PAYMENT_METHODS or amount <= 0 or not description:
                flash("Lütfen zorunlu masraf bilgilerini kontrol edin.", "error")
            else:
                create_database_backup(app, "before_expense_edit")
                expense.expense_date = parse_date(request.form.get("expense_date")) or expense.expense_date
                expense.category = category
                expense.document_no = request.form.get("document_no", "").strip()
                expense.payee = request.form.get("payee", "").strip()
                expense.description = description
                expense.payment_method = payment_method
                expense.amount = amount
                db.session.commit()
                flash("Masraf kaydı güncellendi.", "success")
                return redirect(url_for("expenses"))
        return render_template("expense_edit.html", expense=expense, categories=EXPENSE_CATEGORIES, payment_methods=PAYMENT_METHODS)

    @app.post("/masraflar/<int:expense_id>/sil")
    def delete_expense(expense_id):
        expense = db.get_or_404(Expense, expense_id)
        create_database_backup(app, "before_expense_delete")
        db.session.delete(expense)
        db.session.commit()
        flash("Masraf kaydı silindi.", "success")
        return redirect(url_for("expenses"))

    @app.get("/siparisler")
    def orders():
        query = request.args.get("q", "").strip()
        status = request.args.get("status", "").strip()
        order_type = request.args.get("type", "").strip()
        active_only = request.args.get("active") == "1"
        delivery_pending = request.args.get("delivery_pending") == "1"
        status_summary = {
            kind: {order_status: {"count": 0, "total": Decimal("0")} for order_status in ORDER_STATUSES}
            for kind in ORDER_TYPES
        }
        for summary_order in Order.query.all():
            if summary_order.order_type not in status_summary or summary_order.status not in status_summary[summary_order.order_type]:
                continue
            bucket = status_summary[summary_order.order_type][summary_order.status]
            bucket["count"] += 1
            bucket["total"] += summary_order.total_amount
        records = Order.query.join(Customer)
        if query:
            records = records.filter(db.or_(Order.order_no.ilike(f"%{query}%"), Customer.name.ilike(f"%{query}%")))
        if status:
            records = records.filter(Order.status == status)
        if active_only:
            records = records.filter(~Order.status.in_(["Teslim Edildi", "İptal Edildi"]))
        if delivery_pending:
            records = records.filter(Order.status != "İptal Edildi")
        if order_type in ORDER_TYPES:
            records = records.filter(Order.order_type == order_type)
        if delivery_pending:
            all_cashflow_orders = Order.query.filter(Order.status != "İptal Edildi").all()
            pending_expected = calculate_pending_delivery_amounts(all_cashflow_orders)
            open_ids = {order.id for order in all_cashflow_orders if pending_expected.get(order.id, 0) > 0}
            records = records.filter(Order.id.in_(open_ids)) if open_ids else records.filter(db.false())
            counts = {kind: sum(1 for order in all_cashflow_orders if order.order_type == kind and order.id in open_ids) for kind in ORDER_TYPES}
        elif active_only:
            counts = {kind: Order.query.filter(Order.order_type == kind, ~Order.status.in_(["Teslim Edildi", "İptal Edildi"])).count() for kind in ORDER_TYPES}
        else:
            counts = {kind: Order.query.filter_by(order_type=kind).count() for kind in ORDER_TYPES}
        listed_orders = records.order_by(Order.delivery_date.asc().nullslast(), Order.order_date.desc(), Order.id.desc()).all() if delivery_pending else records.order_by(Order.order_date.desc(), Order.id.desc()).all()
        listed_sales_ids = [order.id for order in listed_orders if order.order_type == "Satış"]
        linked_by_source = {order_id: [] for order_id in listed_sales_ids}
        if listed_sales_ids:
            for linked_order in Order.query.filter(Order.source_order_id.in_(listed_sales_ids)).order_by(Order.id).all():
                linked_by_source[linked_order.source_order_id].append(linked_order)
        procurement_summaries = {
            order.id: procurement_summary(order, linked_by_source.get(order.id, []))
            for order in listed_orders if order.order_type == "Satış"
        }
        pending_expected = pending_expected if delivery_pending else {}
        listed_total = sum((pending_expected.get(order.id, order.total_amount) for order in listed_orders), Decimal("0"))
        return render_template("orders.html", orders=listed_orders, statuses=ORDER_STATUSES, query=query, selected_status=status, selected_type=order_type, active_only=active_only, delivery_pending=delivery_pending, listed_total=listed_total, pending_expected=pending_expected, type_counts=counts, status_summary=status_summary, procurement_summaries=procurement_summaries)

    @app.route("/siparisler/yeni", methods=["GET", "POST"])
    def new_order():
        customers_list = Customer.query.order_by(Customer.name).all()
        products_list = Product.query.filter_by(active=True).order_by(Product.name).all()
        if request.method == "POST":
            customer_id = request.form.get("customer_id", type=int)
            order_type = request.form.get("order_type", "Satış")
            names = request.form.getlist("product_name[]")
            quantities = request.form.getlist("quantity[]")
            if order_type not in ORDER_TYPES:
                flash("Lütfen geçerli bir sipariş türü seçin.", "error")
            elif not customer_id or not db.session.get(Customer, customer_id):
                flash("Lütfen bir müşteri seçin.", "error")
            elif not any(name.strip() for name in names):
                flash("En az bir sipariş kalemi ekleyin.", "error")
            else:
                order = Order(order_no=next_order_no(order_type), order_type=order_type, customer_id=customer_id, order_date=parse_date(request.form.get("order_date")) or date.today(), delivery_date=parse_date(request.form.get("delivery_date")), notes=request.form.get("notes"), status="Bekliyor")
                db.session.add(order)
                product_ids = request.form.getlist("product_id[]")
                variants = request.form.getlist("variant[]")
                details_2 = request.form.getlist("detail_2[]")
                details_3 = request.form.getlist("detail_3[]")
                units = request.form.getlist("unit[]")
                prices = request.form.getlist("unit_price[]")
                discount_rates = request.form.getlist("discount_rate[]")
                vat_rates = request.form.getlist("vat_rate[]")
                item_notes = request.form.getlist("item_note[]")
                for i, name in enumerate(names):
                    if not name.strip():
                        continue
                    quantity = int(quantities[i]) if i < len(quantities) and quantities[i].isdigit() else 1
                    vat_rate = parse_money(vat_rates[i] if i < len(vat_rates) else "10")
                    discount_rate = parse_money(discount_rates[i] if i < len(discount_rates) else "0")
                    product_id = int(product_ids[i]) if i < len(product_ids) and product_ids[i].isdigit() else None
                    order.items.append(OrderItem(product_id=product_id, product_name=name.strip(), description="", variant=variants[i] if i < len(variants) else "", detail_2=details_2[i] if i < len(details_2) else "", detail_3=details_3[i] if i < len(details_3) else "", quantity=max(quantity, 1), unit=units[i] if i < len(units) and units[i] else "Adet", unit_price=parse_money(prices[i] if i < len(prices) else "0"), discount_rate=max(Decimal("0"), min(discount_rate, Decimal("100"))), cost_unit_price=Decimal("0"), vat_rate=max(Decimal("0"), min(vat_rate, Decimal("100"))), note=item_notes[i] if i < len(item_notes) else ""))
                order.history.append(OrderHistory(status="Bekliyor", note="Sipariş oluşturuldu"))
                db.session.commit()
                flash(f"{order.order_no} numaralı sipariş oluşturuldu.", "success")
                return redirect(url_for("order_detail", order_id=order.id))
        return render_template("order_form.html", customers=customers_list, products=products_list, order_types=ORDER_TYPES, selected_type=request.args.get("type", "Satış"), today=date.today().isoformat())

    @app.get("/siparisler/<int:order_id>")
    def order_detail(order_id):
        order = db.get_or_404(Order, order_id)
        converted_orders = Order.query.filter_by(source_order_id=order.id).order_by(Order.id).all() if order.order_type == "Satış" else []
        procurement = procurement_summary(order, converted_orders) if order.order_type == "Satış" else None
        return render_template("order_detail.html", order=order, converted_orders=converted_orders, procurement=procurement, statuses=ORDER_STATUSES)

    @app.get("/siparisler/<int:order_id>/excel")
    def export_order_excel(order_id):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        order = db.get_or_404(Order, order_id)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sipariş"
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A9"
        blue, navy, pale, line = "2563EB", "172033", "EEF4FF", "D8E0EC"
        sheet.merge_cells("A1:K1")
        sheet["A1"] = f"BUSINESS OS - {order.order_type.upper()} SİPARİŞİ"
        sheet["A1"].font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor=navy)
        sheet["A1"].alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 34
        info = [
            ("Sipariş No", order.order_no, "Cari", order.customer.name),
            ("Sipariş Tarihi", order.order_date, "Teslim Tarihi", order.delivery_date or "Belirtilmedi"),
            ("Durum", order.status, "Tür", order.order_type),
        ]
        for row_no, values in enumerate(info, 3):
            sheet.cell(row_no, 1, values[0]); sheet.cell(row_no, 2, values[1])
            sheet.cell(row_no, 6, values[2]); sheet.cell(row_no, 7, values[3])
            sheet.merge_cells(start_row=row_no, start_column=2, end_row=row_no, end_column=5)
            sheet.merge_cells(start_row=row_no, start_column=7, end_row=row_no, end_column=11)
            for col in (1, 6):
                sheet.cell(row_no, col).font = Font(name="Arial", bold=True, color="64748B")
            for col in (2, 7):
                sheet.cell(row_no, col).font = Font(name="Arial", bold=True, color=navy)
        for cell_ref in ("B4", "G4"):
            if hasattr(sheet[cell_ref].value, "year"):
                sheet[cell_ref].number_format = "dd.mm.yyyy"
        headers = ["Sıra", "Ürün", "Ayrıntı 1", "Ayrıntı 2", "Ayrıntı 3", "Adet", "Birim", "Birim Fiyat", "KDV %", "KDV Tutarı", "KDV Dahil Toplam"]
        header_row = 8
        for column, heading in enumerate(headers, 1):
            cell = sheet.cell(header_row, column, heading)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=blue)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[header_row].height = 30
        thin = Side(style="thin", color=line)
        first_data_row = header_row + 1
        for index, item in enumerate(order.items, 1):
            row_no = header_row + index
            values = [index, item.product_name, item.variant or "", item.detail_2 or "", item.detail_3 or "", item.quantity, item.unit, float(item.unit_price or 0), float(item.vat_rate or 0) / 100]
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row_no, column, value)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=column in (2, 3, 4, 5))
                cell.border = Border(bottom=thin)
            sheet.cell(row_no, 10, f"=F{row_no}*H{row_no}*I{row_no}")
            sheet.cell(row_no, 11, f"=F{row_no}*H{row_no}+J{row_no}")
            for column in (8, 10, 11):
                sheet.cell(row_no, column).number_format = '₺#,##0.00'
            sheet.cell(row_no, 9).number_format = "0.00%"
            sheet.row_dimensions[row_no].height = 30
        last_data_row = header_row + len(order.items)
        total_row = last_data_row + 2
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
        sheet.cell(total_row, 8, "Ara Toplam")
        sheet.cell(total_row, 11, f"=SUMPRODUCT(F{first_data_row}:F{last_data_row},H{first_data_row}:H{last_data_row})")
        sheet.cell(total_row + 1, 8, "Toplam KDV")
        sheet.cell(total_row + 1, 11, f"=SUM(J{first_data_row}:J{last_data_row})")
        sheet.cell(total_row + 2, 8, "Genel Toplam")
        sheet.cell(total_row + 2, 11, f"=SUM(K{first_data_row}:K{last_data_row})")
        for row_no in range(total_row, total_row + 3):
            sheet.merge_cells(start_row=row_no, start_column=8, end_row=row_no, end_column=10)
            sheet.cell(row_no, 8).font = Font(name="Arial", bold=True, color=navy)
            sheet.cell(row_no, 11).font = Font(name="Arial", bold=True, color=blue, size=12 if row_no == total_row + 2 else 10)
            sheet.cell(row_no, 11).number_format = '₺#,##0.00'
            sheet.cell(row_no, 8).fill = sheet.cell(row_no, 11).fill = PatternFill("solid", fgColor=pale)
        if order.notes:
            note_row = total_row + 4
            sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=11)
            sheet.cell(note_row, 1, f"Sipariş Notu: {order.notes}")
            sheet.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
            sheet.cell(note_row, 1).fill = PatternFill("solid", fgColor="FFF8E8")
            sheet.row_dimensions[note_row].height = 36
        widths = [7, 30, 18, 18, 18, 10, 10, 15, 10, 15, 18]
        for column, width in enumerate(widths, 1):
            sheet.column_dimensions[chr(64 + column)].width = width
        sheet.auto_filter.ref = f"A8:K{last_data_row}"
        sheet.print_title_rows = "1:8"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"{order.order_no}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.get("/siparisler/<int:order_id>/pdf")
    def export_order_pdf(order_id):
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        order = db.get_or_404(Order, order_id)
        regular_font = "/System/Library/Fonts/Supplemental/Arial.ttf"
        bold_font = "/System/Library/Fonts/Supplemental/Arial Bold.ttf"
        if os.path.exists(regular_font) and os.path.exists(bold_font):
            pdfmetrics.registerFont(TTFont("BusinessArial", regular_font))
            pdfmetrics.registerFont(TTFont("BusinessArialBold", bold_font))
            font_name, bold_name = "BusinessArial", "BusinessArialBold"
        else:
            font_name, bold_name = "Helvetica", "Helvetica-Bold"
        output = BytesIO()
        document = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm, title=order.order_no)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("TitleTR", parent=styles["Title"], fontName=bold_name, fontSize=17, leading=20, textColor=colors.HexColor("#172033"), alignment=TA_LEFT)
        body_style = ParagraphStyle("BodyTR", parent=styles["BodyText"], fontName=font_name, fontSize=7.5, leading=9)
        body_bold = ParagraphStyle("BodyBoldTR", parent=body_style, fontName=bold_name)
        header_style = ParagraphStyle("HeaderTR", parent=body_bold, textColor=colors.white, alignment=TA_LEFT)
        right_style = ParagraphStyle("RightTR", parent=body_style, alignment=TA_RIGHT)
        story = [Paragraph(f"BUSINESS OS - {order.order_type.upper()} SİPARİŞİ", title_style), Spacer(1, 4*mm)]
        info_data = [[Paragraph("Sipariş No", body_bold), Paragraph(order.order_no, body_style), Paragraph("Cari", body_bold), Paragraph(order.customer.name, body_style)], [Paragraph("Sipariş Tarihi", body_bold), Paragraph(order.order_date.strftime("%d.%m.%Y"), body_style), Paragraph("Teslim Tarihi", body_bold), Paragraph(order.delivery_date.strftime("%d.%m.%Y") if order.delivery_date else "Belirtilmedi", body_style)], [Paragraph("Durum", body_bold), Paragraph(order.status, body_style), Paragraph("Tür", body_bold), Paragraph(order.order_type, body_style)]]
        info_table = Table(info_data, colWidths=[28*mm, 70*mm, 28*mm, 125*mm])
        info_table.setStyle(TableStyle([("BACKGROUND",(0,0),(0,-1),colors.HexColor("#EEF4FF")),("BACKGROUND",(2,0),(2,-1),colors.HexColor("#EEF4FF")),("FONTNAME",(0,0),(-1,-1),font_name),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#D8E0EC")),("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
        story.extend([info_table, Spacer(1, 5*mm)])
        headers = ["Sıra", "Ürün", "Ayrıntı 1", "Ayrıntı 2", "Ayrıntı 3", "Adet", "Birim", "Birim Fiyat", "KDV %", "KDV", "Toplam"]
        data = [[Paragraph(value, header_style) for value in headers]]
        money_text = lambda value: f"TL {Decimal(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        for index, item in enumerate(order.items, 1):
            data.append([Paragraph(str(index), body_style), Paragraph(item.product_name, body_style), Paragraph(item.variant or "-", body_style), Paragraph(item.detail_2 or "-", body_style), Paragraph(item.detail_3 or "-", body_style), Paragraph(str(item.quantity), right_style), Paragraph(item.unit, body_style), Paragraph(money_text(item.unit_price or 0), right_style), Paragraph(f"%{item.vat_rate}", right_style), Paragraph(money_text(item.vat_amount), right_style), Paragraph(money_text(item.total_amount), right_style)])
        item_table = Table(data, repeatRows=1, colWidths=[10*mm, 44*mm, 29*mm, 29*mm, 29*mm, 13*mm, 15*mm, 24*mm, 15*mm, 23*mm, 25*mm])
        item_table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2563EB")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,-1),font_name),("VALIGN",(0,0),(-1,-1),"TOP"),("LINEBELOW",(0,0),(-1,-1),0.35,colors.HexColor("#D8E0EC")),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F8FAFC")]),("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
        story.extend([item_table, Spacer(1, 4*mm)])
        totals = [[Paragraph("Ara Toplam", body_bold), Paragraph(money_text(order.net_amount), right_style)], [Paragraph("Toplam KDV", body_bold), Paragraph(money_text(order.vat_amount), right_style)], [Paragraph("KDV Dahil Genel Toplam", body_bold), Paragraph(money_text(order.total_amount), right_style)]]
        totals_table = Table(totals, colWidths=[55*mm, 38*mm], hAlign="RIGHT")
        totals_table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EEF4FF")),("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#D8E0EC")),("FONTNAME",(0,0),(-1,-1),font_name),("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
        story.append(totals_table)
        if order.notes:
            story.extend([Spacer(1, 4*mm), Paragraph(f"Sipariş Notu: {order.notes}", body_bold)])
        def page_number(canvas, doc):
            canvas.saveState(); canvas.setFont(font_name, 7); canvas.setFillColor(colors.HexColor("#64748B")); canvas.drawRightString(landscape(A4)[0]-12*mm, 7*mm, f"Sayfa {doc.page}"); canvas.restoreState()
        document.build(story, onFirstPage=page_number, onLaterPages=page_number)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"{order.order_no}.pdf", mimetype="application/pdf")

    @app.route("/siparisler/<int:order_id>/duzenle", methods=["GET", "POST"])
    def edit_order(order_id):
        order = db.get_or_404(Order, order_id)
        customers_list = Customer.query.order_by(Customer.name).all()
        products_list = Product.query.filter_by(active=True).order_by(Product.name).all()
        if request.method == "POST":
            customer_id = request.form.get("customer_id", type=int)
            names = request.form.getlist("product_name[]")
            quantities = request.form.getlist("quantity[]")
            if not customer_id or not db.session.get(Customer, customer_id):
                flash("Lütfen bir müşteri veya tedarikçi seçin.", "error")
            elif not any(name.strip() for name in names):
                flash("En az bir sipariş kalemi ekleyin.", "error")
            else:
                create_database_backup(app, "before_order_edit")
                order.customer_id = customer_id
                order.order_date = parse_date(request.form.get("order_date")) or order.order_date
                order.delivery_date = parse_date(request.form.get("delivery_date"))
                order.notes = request.form.get("notes", "").strip()
                previous_costs = {(item.product_id, item.product_name): item.cost_unit_price for item in order.items}
                order.items.clear()
                product_ids = request.form.getlist("product_id[]")
                variants = request.form.getlist("variant[]")
                details_2 = request.form.getlist("detail_2[]")
                details_3 = request.form.getlist("detail_3[]")
                units = request.form.getlist("unit[]")
                prices = request.form.getlist("unit_price[]")
                discount_rates = request.form.getlist("discount_rate[]")
                vat_rates = request.form.getlist("vat_rate[]")
                item_notes = request.form.getlist("item_note[]")
                for index, name in enumerate(names):
                    if not name.strip():
                        continue
                    quantity_text = quantities[index] if index < len(quantities) else "1"
                    try:
                        quantity = max(int(quantity_text), 1)
                    except (TypeError, ValueError):
                        quantity = 1
                    product_id = int(product_ids[index]) if index < len(product_ids) and product_ids[index].isdigit() else None
                    vat_rate = parse_money(vat_rates[index] if index < len(vat_rates) else "10")
                    discount_rate = parse_money(discount_rates[index] if index < len(discount_rates) else "0")
                    saved_cost = previous_costs.get((product_id, name.strip()))
                    cost = saved_cost if saved_cost is not None else Decimal("0")
                    order.items.append(OrderItem(product_id=product_id, product_name=name.strip(), description="", variant=variants[index] if index < len(variants) else "", detail_2=details_2[index] if index < len(details_2) else "", detail_3=details_3[index] if index < len(details_3) else "", quantity=quantity, unit=units[index] if index < len(units) and units[index] else "Adet", unit_price=parse_money(prices[index] if index < len(prices) else "0"), discount_rate=max(Decimal("0"), min(discount_rate, Decimal("100"))), cost_unit_price=cost, vat_rate=max(Decimal("0"), min(vat_rate, Decimal("100"))), note=item_notes[index] if index < len(item_notes) else ""))
                order.history.append(OrderHistory(status=order.status, note="Sipariş bilgileri düzenlendi"))
                db.session.commit()
                flash(f"{order.order_no} numaralı sipariş güncellendi.", "success")
                return redirect(url_for("order_detail", order_id=order.id))
        initial_items = [{"product_id": item.product_id, "product_name": item.product_name, "product_label": (f"{item.product.name} · {item.product.code}" if item.product and item.product.code else item.product.name if item.product else ""), "variant": item.variant or "", "detail_2": item.detail_2 or "", "detail_3": item.detail_3 or "", "quantity": item.quantity, "unit": item.unit, "unit_price": str(item.unit_price or 0), "discount_rate": str(item.discount_rate or 0), "vat_rate": str(item.vat_rate or 0), "note": item.note or ""} for item in order.items]
        return render_template("order_form.html", order=order, initial_items=initial_items, customers=customers_list, products=products_list, order_types=ORDER_TYPES, selected_type=order.order_type, today=order.order_date.isoformat())

    @app.route("/siparisler/<int:order_id>/satinalmaya-donustur", methods=["GET", "POST"])
    def convert_to_purchase(order_id):
        source_order = db.get_or_404(Order, order_id)
        if source_order.order_type != "Satış":
            flash("Yalnızca satış siparişleri satın alma siparişine dönüştürülebilir.", "error")
            return redirect(url_for("order_detail", order_id=source_order.id))
        suppliers = Customer.query.order_by(Customer.name).all()
        converted_orders = Order.query.filter_by(source_order_id=source_order.id).order_by(Order.id).all()
        procurement = procurement_summary(source_order, converted_orders)
        if request.method == "POST":
            supplier_id = request.form.get("customer_id", type=int)
            selected_ids = {value for value in request.form.getlist("source_item_id[]") if value.isdigit()}
            selected_items = [item for item in source_order.items if str(item.id) in selected_ids and procurement["items"][item.id]["remaining"] > 0]
            if not supplier_id or not db.session.get(Customer, supplier_id):
                flash("Lütfen bir tedarikçi seçin.", "error")
            elif not selected_items:
                flash("Satın alma siparişine aktarılacak en az bir ürün seçin.", "error")
            else:
                create_database_backup(app, "before_order_conversion")
                purchase = Order(order_no=next_order_no("Satın Alma"), order_type="Satın Alma", source_order_id=source_order.id, customer_id=supplier_id, order_date=parse_date(request.form.get("order_date")) or date.today(), delivery_date=parse_date(request.form.get("delivery_date")), notes=request.form.get("notes", "").strip(), status="Bekliyor")
                for source_item in selected_items:
                    item_id = source_item.id
                    vat_rate = parse_money(request.form.get(f"vat_rate_{item_id}", "10"))
                    remaining_quantity = procurement["items"][source_item.id]["remaining"]
                    quantity = request.form.get(f"quantity_{item_id}", type=int) or remaining_quantity
                    purchase.items.append(OrderItem(source_order_item_id=source_item.id, product_id=source_item.product_id, product_name=source_item.product_name, description="", variant=request.form.get(f"variant_{item_id}", source_item.variant), detail_2=request.form.get(f"detail_2_{item_id}", source_item.detail_2), detail_3=request.form.get(f"detail_3_{item_id}", source_item.detail_3), quantity=max(1, min(quantity, remaining_quantity)), unit=source_item.unit, unit_price=parse_money(request.form.get(f"unit_price_{item_id}", "0")), vat_rate=max(Decimal("0"), min(vat_rate, Decimal("100"))), note=request.form.get(f"item_note_{item_id}", source_item.note)))
                purchase.history.append(OrderHistory(status="Bekliyor", note="Satış siparişindeki ürün detaylarından oluşturuldu"))
                db.session.add(purchase)
                db.session.commit()
                flash(f"{purchase.order_no} numaralı satın alma siparişi oluşturuldu. Satış müşterisi ve satış fiyatları aktarılmadı.", "success")
                return redirect(url_for("order_detail", order_id=purchase.id))
        return render_template("order_convert.html", source_order=source_order, suppliers=suppliers, converted_orders=converted_orders, procurement=procurement, today=date.today().isoformat())

    @app.post("/siparisler/<int:order_id>/durum")
    def update_order_status(order_id):
        order = db.get_or_404(Order, order_id)
        status = request.form.get("status")
        if status not in ORDER_STATUSES:
            flash("Geçersiz sipariş durumu.", "error")
        elif status != order.status:
            order.status = status
            if order.order_type == "Satış" and status in FINANCIAL_ORDER_STATUSES:
                for item in order.items:
                    if not item.cost_unit_price:
                        item.cost_unit_price = latest_delivered_purchase_cost(item)
            order.history.append(OrderHistory(status=status, note=request.form.get("note", "").strip() or "Durum güncellendi"))
            db.session.commit()
            flash("Sipariş durumu güncellendi.", "success")
        return redirect(url_for("order_detail", order_id=order.id))

    @app.route("/sahsi-hesaplar", methods=["GET", "POST"])
    def personal_finance():
        tab = request.values.get("tab", "payments")
        selected_month = request.values.get("month")
        selected_person_id = request.values.get("person_id", type=int)

        def optional_money(field_name):
            raw = request.form.get(field_name, "").strip()
            return parse_money(raw) if raw else None

        if request.method == "POST":
            action = request.form.get("action")
            if action == "save_payments":
                create_database_backup(app, "before_personal_payments")
                for entry_id in request.form.getlist("entry_id"):
                    entry = db.session.get(PersonalPayment, int(entry_id))
                    if not entry:
                        continue
                    entry.name = request.form.get(f"name_{entry.id}", "").strip() or entry.name
                    entry.kind = request.form.get(f"kind_{entry.id}", entry.kind)
                    entry.due_day = request.form.get(f"due_day_{entry.id}", type=int)
                    entry.credit_limit = optional_money(f"credit_limit_{entry.id}")
                    entry.debt = optional_money(f"debt_{entry.id}")
                    entry.minimum_payment = optional_money(f"minimum_payment_{entry.id}")
                    entry.payment = optional_money(f"payment_{entry.id}")
                    entry.available_limit = optional_money(f"available_limit_{entry.id}")
                    entry.remaining_debt = optional_money(f"remaining_debt_{entry.id}") if entry.kind != "card" else None
                db.session.commit()
                flash("Şahsi ödeme tablosu kaydedildi.", "success")
            elif action == "add_payment":
                order = db.session.query(db.func.max(PersonalPayment.sort_order)).filter_by(month=selected_month).scalar() or 0
                db.session.add(PersonalPayment(month=selected_month, kind="other", name="Yeni ödeme", sort_order=order + 1))
                db.session.commit()
            elif action == "delete_payment":
                entry = db.session.get(PersonalPayment, request.form.get("entry_id", type=int))
                if entry:
                    db.session.delete(entry)
                    db.session.commit()
            elif action == "new_month":
                new_month = request.form.get("new_month", "")
                source = db.session.get(PersonalMonth, selected_month)
                if len(new_month) == 7 and not db.session.get(PersonalMonth, new_month):
                    target = PersonalMonth(month=new_month)
                    db.session.add(target)
                    db.session.flush()
                    for entry in source.entries if source else []:
                        db.session.add(PersonalPayment(month=new_month, kind=entry.kind, name=entry.name,
                            credit_limit=entry.credit_limit, due_day=entry.due_day, sort_order=entry.sort_order))
                    db.session.commit()
                    selected_month = new_month
                    flash("Yeni dönem oluşturuldu.", "success")
                else:
                    flash("Dönem oluşturulamadı veya zaten mevcut.", "error")
            elif action == "add_person":
                name = request.form.get("person_name", "").strip()
                if name and not PersonalPerson.query.filter(db.func.normalize_tr(PersonalPerson.name) == normalize_search_text(name)).first():
                    person = PersonalPerson(name=name)
                    db.session.add(person)
                    db.session.commit()
                    selected_person_id = person.id
                    flash(f"{name} kişi hesabı eklendi.", "success")
                else:
                    flash("Kişi adı boş veya zaten mevcut.", "error")
            elif action == "save_ledger":
                create_database_backup(app, "before_personal_ledger")
                for transaction_id in request.form.getlist("transaction_id"):
                    transaction = db.session.get(PersonalLedgerTransaction, int(transaction_id))
                    if not transaction:
                        continue
                    transaction.transaction_date = parse_date(request.form.get(f"transaction_date_{transaction.id}"))
                    transaction.description = request.form.get(f"description_{transaction.id}", "").strip()
                    transaction.sent_amount = optional_money(f"sent_amount_{transaction.id}")
                    transaction.received_amount = optional_money(f"received_amount_{transaction.id}")
                    if transaction.sent_amount:
                        transaction.received_amount = None
                new_dates = request.form.getlist("new_transaction_date[]")
                new_descriptions = request.form.getlist("new_description[]")
                new_sent = request.form.getlist("new_sent_amount[]")
                new_received = request.form.getlist("new_received_amount[]")
                order = db.session.query(db.func.max(PersonalLedgerTransaction.sort_order)).filter_by(person_id=selected_person_id).scalar() or 0
                for index, description in enumerate(new_descriptions):
                    sent = parse_money(new_sent[index]) if index < len(new_sent) and new_sent[index].strip() else None
                    received = parse_money(new_received[index]) if index < len(new_received) and new_received[index].strip() else None
                    if not description.strip() and not sent and not received:
                        continue
                    order += 1
                    db.session.add(PersonalLedgerTransaction(person_id=selected_person_id,
                        transaction_date=parse_date(new_dates[index]) if index < len(new_dates) else date.today(),
                        description=description.strip(), sent_amount=sent, received_amount=None if sent else received, sort_order=order))
                db.session.commit()
                flash("Borç–alacak hesabı kaydedildi.", "success")
            elif action == "delete_transaction":
                transaction = db.session.get(PersonalLedgerTransaction, request.form.get("transaction_id", type=int))
                if transaction:
                    selected_person_id = transaction.person_id
                    db.session.delete(transaction)
                    db.session.commit()
            return redirect(url_for("personal_finance", tab=tab, month=selected_month, person_id=selected_person_id))

        months = PersonalMonth.query.order_by(PersonalMonth.month.desc()).all()
        if not months:
            selected_month = date.today().strftime("%Y-%m")
            db.session.add(PersonalMonth(month=selected_month))
            db.session.commit()
            months = PersonalMonth.query.all()
        if not selected_month or not db.session.get(PersonalMonth, selected_month):
            selected_month = months[0].month
        entries = PersonalPayment.query.filter_by(month=selected_month).order_by(PersonalPayment.sort_order, PersonalPayment.id).all()
        payment_totals = {
            "debt": sum((item.debt or Decimal("0") for item in entries), Decimal("0")),
            "minimum": sum((item.minimum_payment or Decimal("0") for item in entries), Decimal("0")),
            "payment": sum((item.payment or Decimal("0") for item in entries), Decimal("0")),
            "available_limit": sum((item.available_limit or Decimal("0") for item in entries if item.kind == "card"), Decimal("0")),
            "remaining": sum((item.calculated_remaining for item in entries), Decimal("0")),
        }
        people = PersonalPerson.query.order_by(PersonalPerson.name).all()
        person = db.session.get(PersonalPerson, selected_person_id) if selected_person_id else (people[0] if people else None)
        transactions = list(person.transactions) if person else []
        sent_total = sum((item.sent_amount or Decimal("0") for item in transactions), Decimal("0"))
        received_total = sum((item.received_amount or Decimal("0") for item in transactions), Decimal("0"))
        return render_template("personal_finance.html", tab=tab, months=months, selected_month=selected_month,
            entries=entries, payment_totals=payment_totals, people=people, person=person, transactions=transactions,
            sent_total=sent_total, received_total=received_total, balance=received_total-sent_total, today=date.today().isoformat())

    @app.post("/sahsi-hesaplar/odeme/<int:entry_id>/sil")
    def delete_personal_payment(entry_id):
        entry = db.get_or_404(PersonalPayment, entry_id)
        month = entry.month
        db.session.delete(entry)
        db.session.commit()
        flash("Ödeme satırı silindi.", "success")
        return redirect(url_for("personal_finance", tab="payments", month=month))

    @app.post("/sahsi-hesaplar/hareket/<int:transaction_id>/sil")
    def delete_personal_transaction(transaction_id):
        transaction = db.get_or_404(PersonalLedgerTransaction, transaction_id)
        person_id = transaction.person_id
        db.session.delete(transaction)
        db.session.commit()
        flash("Hareket silindi.", "success")
        return redirect(url_for("personal_finance", tab="ledger", person_id=person_id))

    @app.get("/sahsi-hesaplar/borc-alacak/<int:person_id>/pdf")
    def export_personal_ledger_pdf(person_id):
        person = db.get_or_404(PersonalPerson, person_id)
        filename = f"{safe_export_name(person.name)}-Borc-Alacak-Ekstresi.pdf"
        return send_file(build_personal_ledger_pdf(person), mimetype="application/pdf", as_attachment=True, download_name=filename)

    @app.get("/sahsi-hesaplar/borc-alacak/<int:person_id>/excel")
    def export_personal_ledger_xlsx(person_id):
        person = db.get_or_404(PersonalPerson, person_id)
        filename = f"{safe_export_name(person.name)}-Borc-Alacak-Ekstresi.xlsx"
        return send_file(build_personal_ledger_xlsx(person),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)

    @app.cli.command("init-db")
    def init_db_command():
        db.create_all()
        print("Veritabanı hazırlandı.")

    with app.app_context():
        create_database_backup(app, "startup")
        db.create_all()
        import_personal_finance_data(app)
        # Küçük SQLite kurulumlarında ayrıca bir migration aracı gerektirmeden
        # eski müşteri tablolarını yeni alanlarla uyumlu hale getirir.
        if db.engine.dialect.name == "sqlite":
            customer_columns = {column["name"] for column in inspect(db.engine).get_columns("customer")}
            if "code" not in customer_columns:
                db.session.execute(text("ALTER TABLE customer ADD COLUMN code VARCHAR(80)"))
                db.session.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_customer_code ON customer (code)"))
            if "mobile" not in customer_columns:
                db.session.execute(text("ALTER TABLE customer ADD COLUMN mobile VARCHAR(40)"))
            order_columns = {column["name"] for column in inspect(db.engine).get_columns("order")}
            if "order_type" not in order_columns:
                db.session.execute(text("ALTER TABLE 'order' ADD COLUMN order_type VARCHAR(30) NOT NULL DEFAULT 'Satış'"))
                db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_order_order_type ON 'order' (order_type)"))
            product_columns = {column["name"] for column in inspect(db.engine).get_columns("product")}
            if "special_code" not in product_columns:
                db.session.execute(text("ALTER TABLE product ADD COLUMN special_code VARCHAR(160)"))
            if "group_name" not in product_columns:
                db.session.execute(text("ALTER TABLE product ADD COLUMN group_name VARCHAR(160)"))
                db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_product_group_name ON product (group_name)"))
            if "purchase_price" not in product_columns:
                db.session.execute(text("ALTER TABLE product ADD COLUMN purchase_price NUMERIC(12, 2) NOT NULL DEFAULT 0"))
            if "include_in_catalog" not in product_columns:
                db.session.execute(text("ALTER TABLE product ADD COLUMN include_in_catalog BOOLEAN NOT NULL DEFAULT 0"))
                db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_product_include_in_catalog ON product (include_in_catalog)"))
            if "include_in_price_list" not in product_columns:
                db.session.execute(text("ALTER TABLE product ADD COLUMN include_in_price_list BOOLEAN NOT NULL DEFAULT 0"))
                db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_product_include_in_price_list ON product (include_in_price_list)"))
            if "source_order_id" not in order_columns:
                db.session.execute(text("ALTER TABLE 'order' ADD COLUMN source_order_id INTEGER"))
            source_index = next((item for item in inspect(db.engine).get_indexes("order") if item["name"] == "ix_order_source_order_id"), None)
            if source_index and source_index.get("unique"):
                db.session.execute(text("DROP INDEX ix_order_source_order_id"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_order_source_order_id ON 'order' (source_order_id)"))
            item_columns = {column["name"] for column in inspect(db.engine).get_columns("order_item")}
            if "detail_2" not in item_columns:
                db.session.execute(text("ALTER TABLE order_item ADD COLUMN detail_2 VARCHAR(160)"))
            if "detail_3" not in item_columns:
                db.session.execute(text("ALTER TABLE order_item ADD COLUMN detail_3 VARCHAR(160)"))
            if "vat_rate" not in item_columns:
                # Eski siparişlerin toplamını değiştirmemek için geçmiş kalemlerde KDV %0 kalır.
                db.session.execute(text("ALTER TABLE order_item ADD COLUMN vat_rate NUMERIC(5, 2) NOT NULL DEFAULT 0"))
            if "discount_rate" not in item_columns:
                db.session.execute(text("ALTER TABLE order_item ADD COLUMN discount_rate NUMERIC(5, 2) NOT NULL DEFAULT 0"))
            if "cost_unit_price" not in item_columns:
                db.session.execute(text("ALTER TABLE order_item ADD COLUMN cost_unit_price NUMERIC(12, 2) NOT NULL DEFAULT 0"))
                # Mevcut satış kalemleri için ürün kartındaki güncel alış fiyatını
                # bir defaya mahsus başlangıç maliyeti olarak sabitler.
                db.session.execute(text("""
                    UPDATE order_item
                    SET cost_unit_price = COALESCE((
                        SELECT product.purchase_price FROM product
                        WHERE product.id = order_item.product_id
                    ), 0)
                    WHERE order_id IN (SELECT id FROM 'order' WHERE order_type = 'Satış')
                """))
            if "source_order_item_id" not in item_columns:
                db.session.execute(text("ALTER TABLE order_item ADD COLUMN source_order_item_id INTEGER"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_order_item_source_order_item_id ON order_item (source_order_item_id)"))
            account_columns = {column["name"] for column in inspect(db.engine).get_columns("account_transaction")}
            if "payment_method" not in account_columns:
                db.session.execute(text("ALTER TABLE account_transaction ADD COLUMN payment_method VARCHAR(30)"))
            if "check_no" not in account_columns:
                db.session.execute(text("ALTER TABLE account_transaction ADD COLUMN check_no VARCHAR(80)"))
            if "check_bank" not in account_columns:
                db.session.execute(text("ALTER TABLE account_transaction ADD COLUMN check_bank VARCHAR(120)"))
            if "check_due_date" not in account_columns:
                db.session.execute(text("ALTER TABLE account_transaction ADD COLUMN check_due_date DATE"))
                db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_account_transaction_check_due_date ON account_transaction (check_due_date)"))
            if "check_status" not in account_columns:
                db.session.execute(text("ALTER TABLE account_transaction ADD COLUMN check_status VARCHAR(40)"))
            if "card_installments" not in account_columns:
                db.session.execute(text("ALTER TABLE account_transaction ADD COLUMN card_installments INTEGER"))
            if "card_owner_type" not in account_columns:
                db.session.execute(text("ALTER TABLE account_transaction ADD COLUMN card_owner_type VARCHAR(40)"))
            if "card_customer_id" not in account_columns:
                db.session.execute(text("ALTER TABLE account_transaction ADD COLUMN card_customer_id INTEGER"))
            if "card_customer_name" not in account_columns:
                db.session.execute(text("ALTER TABLE account_transaction ADD COLUMN card_customer_name VARCHAR(160)"))
            db.session.commit()
    return app


app = create_app()


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
